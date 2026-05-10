"""
Flask backend для системы управления клиентской базой
компании FSN GALLERY.
Flask также раздаёт статические файлы frontend.
"""

from flask import Flask, request, jsonify, session, send_from_directory, redirect
from flask_cors import CORS
import pyodbc
import hashlib
import os
import uuid
import io
from datetime import datetime, date
from decimal import Decimal
from functools import wraps
from werkzeug.utils import secure_filename

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Путь к папке frontend (рядом с backend/)
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(BASE_DIR, '..', 'frontend')
UPLOAD_DIR   = os.path.join(BASE_DIR, '..', 'uploads')

# Создаём папку для загрузок если нет
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'jpg', 'jpeg', 'png', 'zip', 'rar'}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB

app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', 'fsn-gallery-secret-key-2026')

# CORS не нужен когда frontend и backend на одном origin,
# но оставим для совместимости
CORS(app, supports_credentials=True)

# ─── Конфигурация подключения к SQL Server ───────────────────────────────────
DB_CONFIG = {
    'server':   os.environ.get('DB_SERVER',   'localhost'),
    'database': os.environ.get('DB_NAME',     'TileCompanyDB'),
    'username': os.environ.get('DB_USER',     ''),
    'password': os.environ.get('DB_PASSWORD', ''),
    'trusted':  os.environ.get('DB_TRUSTED',  'yes'),   # Windows Auth по умолчанию
}


def get_connection():
    if DB_CONFIG['trusted'].lower() == 'yes':
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={DB_CONFIG['server']};"
            f"DATABASE={DB_CONFIG['database']};"
            f"Trusted_Connection=yes;"
        )
    else:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={DB_CONFIG['server']};"
            f"DATABASE={DB_CONFIG['database']};"
            f"UID={DB_CONFIG['username']};"
            f"PWD={DB_CONFIG['password']};"
        )
    return pyodbc.connect(conn_str)


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode('utf-8')).hexdigest()


def row_to_dict(cursor, row):
    """Конвертирует строку pyodbc в словарь, обеспечивая JSON-совместимость."""
    columns = [col[0] for col in cursor.description]
    result = {}
    for col, val in zip(columns, row):
        if isinstance(val, (datetime, date)):
            result[col] = val.isoformat()
        elif isinstance(val, bytes):
            result[col] = val.decode('utf-8')
        elif isinstance(val, Decimal):
            result[col] = float(val)
        else:
            result[col] = val
    return result


def rows_to_list(cursor, rows):
    return [row_to_dict(cursor, r) for r in rows]


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── Раздача frontend ─────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Корень → страница логина."""
    return send_from_directory(FRONTEND_DIR, 'login.html')


@app.route('/<path:filename>')
def static_files(filename):
    """Раздаём любые файлы из папки frontend или uploads."""
    if filename.startswith('uploads/'):
        file_path = os.path.join(BASE_DIR, '..', filename)
        dir_path = os.path.dirname(file_path)
        fname = os.path.basename(file_path)
        return send_from_directory(dir_path, fname)
    return send_from_directory(FRONTEND_DIR, filename)


# ─── Декораторы авторизации ───────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Необходима авторизация'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Необходима авторизация'}), 401
        if session.get('role') != 'admin':
            return jsonify({'error': 'Недостаточно прав'}), 403
        return f(*args, **kwargs)
    return decorated


def manager_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Необходима авторизация'}), 401
        if session.get('role') not in ('manager', 'admin'):
            return jsonify({'error': 'Недостаточно прав'}), 403
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Самостоятельная регистрация менеджера."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Нет данных'}), 400

    login    = (data.get('login') or '').strip()
    password = data.get('password') or ''
    fullname = (data.get('full_name') or '').strip()

    if not login or not password or not fullname:
        return jsonify({'error': 'Логин, пароль и ФИО обязательны'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Пароль должен быть не менее 6 символов'}), 400
    if len(login) < 3:
        return jsonify({'error': 'Логин должен быть не менее 3 символов'}), 400

    pwd_hash = hash_password(password)
    try:
        conn = get_connection()
        cur = conn.cursor()
        # Создаём пользователя с RoleID=NULL и IsActive=0 (ждёт назначения роли)
        cur.execute("""
            INSERT INTO Users (RoleID, Login, PasswordHash, FullName, IsActive, IsPendingRole)
            OUTPUT INSERTED.UserID
            VALUES (NULL, ?, ?, ?, 0, 1)
        """, login, pwd_hash, fullname)
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'user_id': row[0], 'message': 'Регистрация успешна. Дождитесь назначения роли администратором.'}), 201
    except pyodbc.IntegrityError:
        return jsonify({'error': 'Логин уже занят'}), 409
    except Exception as e:
        return jsonify({'error': f'Ошибка: {str(e)}'}), 500


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    if not data or not data.get('login') or not data.get('password'):
        return jsonify({'error': 'Введите логин и пароль'}), 400

    pwd_hash = hash_password(data['password'])
    try:
        conn = get_connection()
        cur = conn.cursor()
        # Проверяем: существует ли пользователь с таким логином/паролем
        cur.execute("""
            SELECT u.UserID, u.Login, u.FullName, u.Email, u.Phone,
                   u.IsActive, u.IsPendingRole, r.RoleName
            FROM Users u
            LEFT JOIN Roles r ON u.RoleID = r.RoleID
            WHERE u.Login = ? AND u.PasswordHash = ?
        """, data['login'], pwd_hash)
        row = cur.fetchone()
        conn.close()
    except Exception as e:
        return jsonify({'error': f'Ошибка подключения к БД: {str(e)}'}), 500

    if not row:
        return jsonify({'error': 'Неверный логин или пароль'}), 401

    user = row_to_dict(cur, row)

    # Пользователь ждёт назначения роли
    if user.get('IsPendingRole') == True or user.get('IsPendingRole') == 1:
        return jsonify({'error': 'Ваш аккаунт ожидает назначения роли администратором. Пожалуйста, обратитесь к администратору.', 'pending': True}), 403

    # Аккаунт заблокирован
    if not user.get('IsActive'):
        return jsonify({'error': 'Ваш аккаунт заблокирован. Обратитесь к администратору.', 'blocked': True}), 403

    session['user_id']  = user['UserID']
    session['login']    = user['Login']
    session['fullname'] = user['FullName']
    session['role']     = user['RoleName']

    # Обновляем LastLoginAt
    try:
        conn_ll = get_connection()
        cur_ll  = conn_ll.cursor()
        cur_ll.execute("UPDATE Users SET LastLoginAt = GETDATE() WHERE UserID = ?", user['UserID'])
        conn_ll.commit()
        conn_ll.close()
    except Exception:
        pass

    return jsonify({
        'user_id':  user['UserID'],
        'login':    user['Login'],
        'fullname': user['FullName'],
        'role':     user['RoleName'],
        'email':    user['Email'],
        'phone':    user['Phone'],
    })


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'message': 'Выход выполнен'})


@app.route('/api/auth/me', methods=['GET'])
@login_required
def me():
    return jsonify({
        'user_id':  session['user_id'],
        'login':    session['login'],
        'fullname': session['fullname'],
        'role':     session['role'],
    })


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD / STATS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/stats/manager', methods=['GET'])
@manager_required
def manager_stats():
    manager_id = session['user_id']
    now = datetime.now()
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Количество клиентов менеджера
        cur.execute("SELECT COUNT(*) FROM Clients WHERE ManagerID=?", manager_id)
        total_clients = cur.fetchone()[0] or 0

        # Продажи за текущий месяц: только завершённые сделки (IsCompleted=1), CompletedAt в этом месяце
        cur.execute("""
            SELECT ISNULL(SUM(ISNULL(d.Budget,0)),0)
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ManagerID = ?
              AND ds.IsCompleted = 1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt)  = ?
              AND MONTH(d.CompletedAt) = ?
        """, manager_id, now.year, now.month)
        monthly_sales = float(cur.fetchone()[0] or 0)

        # Количество завершённых сделок за этот месяц
        cur.execute("""
            SELECT COUNT(*)
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ManagerID = ?
              AND ds.IsCompleted = 1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt)  = ?
              AND MONTH(d.CompletedAt) = ?
        """, manager_id, now.year, now.month)
        monthly_deals = int(cur.fetchone()[0] or 0)

        # Активные сделки (не завершены, не архив)
        cur.execute("""
            SELECT COUNT(*) FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ManagerID = ?
              AND d.IsArchived = 0
              AND ds.IsCompleted = 0
        """, manager_id)
        active_deals = int(cur.fetchone()[0] or 0)

        conn.close()
        return jsonify({
            'TotalClients':  total_clients,
            'MonthlySales':  monthly_sales,
            'MonthlyDeals':  monthly_deals,
            'ActiveDeals':   active_deals,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/company', methods=['GET'])
@admin_required
def company_stats():
    now = datetime.now()
    cur_year  = now.year
    cur_month = now.month
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM Clients")
        total_clients = int(cur.fetchone()[0] or 0)

        # Общая сумма всех завершённых сделок (IsCompleted=1)
        cur.execute("""
            SELECT ISNULL(SUM(ISNULL(d.Budget,0)),0)
            FROM Deals d INNER JOIN DealStages ds ON d.StageID=ds.StageID
            WHERE ds.IsCompleted=1
        """)
        total_sales = float(cur.fetchone()[0] or 0)

        cur.execute("SELECT COUNT(*) FROM Deals")
        total_deals = int(cur.fetchone()[0] or 0)

        # Продажи за месяц: завершённые сделки с CompletedAt в этом месяце
        cur.execute("""
            SELECT ISNULL(SUM(ISNULL(d.Budget,0)),0)
            FROM Deals d INNER JOIN DealStages ds ON d.StageID=ds.StageID
            WHERE ds.IsCompleted=1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt)=? AND MONTH(d.CompletedAt)=?
        """, cur_year, cur_month)
        monthly_sales = float(cur.fetchone()[0] or 0)

        # Количество завершённых сделок за этот месяц
        cur.execute("""
            SELECT COUNT(*)
            FROM Deals d INNER JOIN DealStages ds ON d.StageID=ds.StageID
            WHERE ds.IsCompleted=1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt)=? AND MONTH(d.CompletedAt)=?
        """, cur_year, cur_month)
        monthly_deals = int(cur.fetchone()[0] or 0)

        cur.execute("""
            SELECT COUNT(*) FROM Users
            WHERE RoleID=(SELECT RoleID FROM Roles WHERE RoleName='manager')
              AND IsActive=1
        """)
        active_managers = int(cur.fetchone()[0] or 0)

        cur.execute("""
            SELECT COUNT(*) FROM Clients
            WHERE YEAR(CreatedAt)=? AND MONTH(CreatedAt)=?
        """, cur_year, cur_month)
        new_clients = int(cur.fetchone()[0] or 0)

        cur.execute("SELECT COUNT(*) FROM Users WHERE IsPendingRole=1")
        pending_managers = int(cur.fetchone()[0] or 0)

        conn.close()
        return jsonify({
            'TotalClients':    total_clients,
            'TotalSales':      total_sales,
            'TotalDeals':      total_deals,
            'MonthlySales':    monthly_sales,
            'MonthlyDeals':    monthly_deals,
            'ActiveManagers':  active_managers,
            'NewClients':      new_clients,
            'PendingManagers': pending_managers,
            'CurrentYear':     cur_year,
            'CurrentMonth':    cur_month,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/clients', methods=['GET'])
@login_required
def get_clients():
    search    = request.args.get('search')
    status_id = request.args.get('status_id')
    manager_id_filter = request.args.get('manager_id')

    try:
        conn = get_connection()
        cur = conn.cursor()

        if session['role'] == 'admin':
            cur.execute(
                "EXEC sp_GetAllClients @Search=?, @StatusID=?, @ManagerID=?",
                search, status_id, manager_id_filter
            )
        else:
            cur.execute(
                "EXEC sp_GetManagerClients @ManagerID=?, @Search=?, @StatusID=?",
                session['user_id'], search, status_id
            )

        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>', methods=['GET'])
@login_required
def get_client(client_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                c.ClientID, c.FullName, c.CompanyName, c.Phone, c.Email, c.Address,
                c.BirthDate, c.FirstContactDate, c.TotalPurchases, c.LastPurchaseDate, c.Notes,
                c.ManagerID,
                ct.TypeName   AS ClientType,  ct.ClientTypeID,
                cs.StatusName AS Status,       cs.StatusID,
                ls.SourceName AS LeadSource,   ls.LeadSourceID,
                u.FullName    AS ManagerName,
                c.CreatedAt,  c.UpdatedAt
            FROM Clients c
            INNER JOIN ClientTypes    ct ON c.ClientTypeID = ct.ClientTypeID
            INNER JOIN ClientStatuses cs ON c.StatusID     = cs.StatusID
            INNER JOIN LeadSources    ls ON c.LeadSourceID = ls.LeadSourceID
            INNER JOIN Users          u  ON c.ManagerID    = u.UserID
            WHERE c.ClientID = ?
        """, client_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Клиент не найден'}), 404

        client = row_to_dict(cur, row)
        if session['role'] == 'manager' and client['ManagerID'] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа к этому клиенту'}), 403

        # Заметки
        cur.execute("""
            SELECT cn.NoteID, cn.NoteText, cn.CreatedAt, u.FullName AS AuthorName
            FROM ClientNotes cn
            INNER JOIN Users u ON cn.AuthorID = u.UserID
            WHERE cn.ClientID = ?
            ORDER BY cn.CreatedAt DESC
        """, client_id)
        notes = rows_to_list(cur, cur.fetchall())

        # Сделки
        cur.execute("""
            SELECT d.DealID, d.Title, d.DealType, d.Budget, d.Priority, d.DealDate,
                   d.Deadline, d.Notes AS DealNotes, d.IsArchived,
                   ds.StageName, ds.StageOrder, ds.IsCompleted,
                   u.FullName AS ManagerName
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            INNER JOIN Users u ON d.ManagerID = u.UserID
            WHERE d.ClientID = ?
            ORDER BY d.CreatedAt DESC
        """, client_id)
        deals = rows_to_list(cur, cur.fetchall())

        # Лента событий (включая авто-события из сделок)
        cur.execute("""
            SELECT ce.EventID, ce.EventType, ce.EventDate, ce.Title,
                   ce.Description, ce.CreatedAt, ce.IsAutomatic,
                   ce.SourceDealID, ce.SourceDealTaskID,
                   u.FullName AS AuthorName
            FROM ClientEvents ce
            INNER JOIN Users u ON ce.AuthorID = u.UserID
            WHERE ce.ClientID = ?
            ORDER BY ce.EventDate DESC
        """, client_id)
        events = rows_to_list(cur, cur.fetchall())

        # Документы клиента
        cur.execute("""
            SELECT cd.DocumentID AS DocID, cd.FileName, cd.OriginalName,
                   cd.FileSize, cd.MimeType, cd.UploadedAt,
                   u.FullName AS UploaderName,
                   'client' AS DocSource, NULL AS DealID, NULL AS DealTitle
            FROM ClientDocuments cd
            INNER JOIN Users u ON cd.UploadedBy = u.UserID
            WHERE cd.ClientID = ?
            ORDER BY cd.UploadedAt DESC
        """, client_id)
        client_docs = rows_to_list(cur, cur.fetchall())

        # Документы из сделок клиента
        cur.execute("""
            SELECT dd.DocID, dd.FileName, dd.OriginalName,
                   dd.FileSize, dd.MimeType, dd.UploadedAt,
                   u.FullName AS UploaderName,
                   'deal' AS DocSource, d.DealID, d.Title AS DealTitle
            FROM DealDocuments dd
            INNER JOIN Deals d ON dd.DealID = d.DealID
            INNER JOIN Users u ON dd.UploadedBy = u.UserID
            WHERE d.ClientID = ?
            ORDER BY dd.UploadedAt DESC
        """, client_id)
        deal_docs = rows_to_list(cur, cur.fetchall())

        documents = client_docs + deal_docs
        documents.sort(key=lambda x: x.get('UploadedAt') or '', reverse=True)

        # Задачи клиента из таблицы Tasks
        cur.execute("""
            SELECT t.TaskID, t.TaskType, t.Title, t.Description,
                   t.ScheduledAt, t.CompletedAt, t.IsCompleted, t.CreatedAt,
                   t.DealID, 'client' AS TaskSource, NULL AS DealTitle
            FROM Tasks t
            WHERE t.ClientID = ?
            ORDER BY t.ScheduledAt ASC
        """, client_id)
        client_tasks_list = rows_to_list(cur, cur.fetchall())

        # Задачи из сделок клиента
        cur.execute("""
            SELECT dt.TaskID, dt.TaskType, dt.Title, dt.Description,
                   dt.ScheduledAt, dt.CompletedAt, dt.IsCompleted, dt.CreatedAt,
                   dt.DealID, 'deal' AS TaskSource, d.Title AS DealTitle
            FROM DealTasks dt
            INNER JOIN Deals d ON dt.DealID = d.DealID
            WHERE d.ClientID = ?
            ORDER BY dt.ScheduledAt ASC
        """, client_id)
        deal_tasks_list = rows_to_list(cur, cur.fetchall())

        # Средний чек (по завершённым сделкам с бюджетом)
        cur.execute("""
            SELECT COUNT(*) AS DealsCount,
                   ISNULL(SUM(d.Budget), 0) AS TotalBudget
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ClientID = ? AND ds.IsCompleted = 1 AND d.Budget IS NOT NULL
        """, client_id)
        avg_row = cur.fetchone()
        deals_count  = avg_row[0] if avg_row else 0
        total_budget = float(avg_row[1]) if avg_row else 0.0
        avg_check = (total_budget / deals_count) if deals_count > 0 else 0.0

        conn.close()
        client['notes']        = notes
        client['deals']        = deals
        client['events']       = events
        client['documents']    = documents
        client['client_tasks'] = client_tasks_list
        client['deal_tasks']   = deal_tasks_list
        client['AvgCheck']     = avg_check
        client['TotalPurchases'] = total_budget
        return jsonify(client)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients', methods=['POST'])
@manager_required
def create_client():
    data = request.get_json()
    required = ['full_name', 'phone', 'client_type_id', 'lead_source_id']
    for f in required:
        if not data.get(f):
            return jsonify({'error': f'Поле {f} обязательно'}), 400

    manager_id = session['user_id']
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO Clients
                (ManagerID, ClientTypeID, LeadSourceID, StatusID,
                 FullName, CompanyName, Phone, Email, Address, BirthDate, FirstContactDate, Notes)
            OUTPUT INSERTED.ClientID
            VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            manager_id,
            data['client_type_id'],
            data['lead_source_id'],
            data['full_name'],
            data.get('company_name'),
            data['phone'],
            data.get('email'),
            data.get('address'),
            data.get('birth_date') or None,
            data.get('first_contact_date', date.today().isoformat()),
            data.get('notes')
        )
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'client_id': row[0], 'message': 'Клиент создан'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>', methods=['PUT'])
@login_required
def update_client(client_id):
    data = request.get_json()
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT ManagerID FROM Clients WHERE ClientID=?", client_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Клиент не найден'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        cur.execute("""
            UPDATE Clients SET
                FullName         = ISNULL(?, FullName),
                CompanyName      = ?,
                Phone            = ISNULL(?, Phone),
                Email            = ?,
                Address          = ?,
                BirthDate        = ?,
                ClientTypeID     = ISNULL(?, ClientTypeID),
                LeadSourceID     = ISNULL(?, LeadSourceID),
                StatusID         = ISNULL(?, StatusID),
                FirstContactDate = ISNULL(?, FirstContactDate),
                Notes            = ?
            WHERE ClientID = ?
        """,
            data.get('full_name'),
            data.get('company_name'),
            data.get('phone'),
            data.get('email'),
            data.get('address'),
            data.get('birth_date') or None,
            data.get('client_type_id'),
            data.get('lead_source_id'),
            data.get('status_id'),
            data.get('first_contact_date'),
            data.get('notes'),
            client_id
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Клиент обновлён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>/reassign', methods=['POST'])
@admin_required
def reassign_client(client_id):
    data = request.get_json()
    new_manager_id = data.get('manager_id')
    if not new_manager_id:
        return jsonify({'error': 'Укажите нового менеджера'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("EXEC sp_ReassignClient @ClientID=?, @NewManagerID=?",
                    client_id, new_manager_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Клиент перераспределён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Заметки ─────────────────────────────────────────────────────────────────

@app.route('/api/clients/<int:client_id>/notes', methods=['POST'])
@login_required
def add_note(client_id):
    data = request.get_json()
    if not data.get('note_text'):
        return jsonify({'error': 'Текст заметки обязателен'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO ClientNotes (ClientID, AuthorID, NoteText)
            OUTPUT INSERTED.NoteID, INSERTED.CreatedAt
            VALUES (?, ?, ?)
        """, client_id, session['user_id'], data['note_text'])
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'note_id': row[0], 'created_at': row[1].isoformat()}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/notes/<int:note_id>', methods=['DELETE'])
@login_required
def delete_note(note_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT AuthorID FROM ClientNotes WHERE NoteID=?", note_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Заметка не найдена'}), 404
        if session['role'] != 'admin' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM ClientNotes WHERE NoteID=?", note_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Заметка удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Лента событий ──────────────────────────────────────────────────────────

@app.route('/api/clients/<int:client_id>/events', methods=['GET'])
@login_required
def get_client_events(client_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT ce.EventID, ce.EventType, ce.EventDate, ce.Title,
                   ce.Description, ce.CreatedAt, ce.IsAutomatic,
                   ce.SourceDealID, ce.SourceDealTaskID,
                   u.FullName AS AuthorName
            FROM ClientEvents ce
            INNER JOIN Users u ON ce.AuthorID = u.UserID
            WHERE ce.ClientID = ?
            ORDER BY ce.EventDate DESC
        """, client_id)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>/events', methods=['POST'])
@login_required
def add_client_event(client_id):
    data = request.get_json()
    if not data.get('event_type') or not data.get('title'):
        return jsonify({'error': 'Тип и заголовок события обязательны'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        raw_date = data.get('event_date') or datetime.now().isoformat()
        try:
            raw_date = raw_date.replace('T', ' ')
            if len(raw_date) == 16:
                raw_date += ':00'
            event_date = datetime.strptime(raw_date, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            event_date = datetime.now()
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic)
            OUTPUT INSERTED.EventID, INSERTED.CreatedAt
            VALUES (?, ?, ?, ?, ?, ?, 0)
        """, client_id, session['user_id'],
             data['event_type'], event_date,
             data['title'],
             data.get('description'))
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'event_id': row[0], 'created_at': row[1].isoformat()}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/events/<int:event_id>', methods=['DELETE'])
@login_required
def delete_event(event_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT AuthorID, IsAutomatic FROM ClientEvents WHERE EventID=?", event_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Событие не найдено'}), 404
        # Менеджер не может удалять автоматические события
        if session['role'] == 'manager' and (row[1] == 1 or row[1] == True):
            conn.close()
            return jsonify({'error': 'Автоматические события нельзя удалять'}), 403
        if session['role'] != 'admin' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM ClientEvents WHERE EventID=?", event_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Событие удалено'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Задачи / планирование ───────────────────────────────────────────────────

@app.route('/api/tasks', methods=['GET'])
@login_required
def get_tasks():
    manager_id = session['user_id']
    only_pending = request.args.get('pending')  # ?pending=1 — только незавершённые
    client_id_filter = request.args.get('client_id')
    try:
        conn = get_connection()
        cur = conn.cursor()
        # Задачи клиента (включая задачи из сделок клиента)
        query = """
            SELECT t.TaskID, t.TaskType, t.Title, t.Description,
                   t.ScheduledAt, t.CompletedAt, t.IsCompleted, t.CreatedAt,
                   c.FullName AS ClientName, c.ClientID, c.Phone AS ClientPhone,
                   t.DealID, 'client' AS TaskSource,
                   NULL AS DealTitle
            FROM Tasks t
            INNER JOIN Clients c ON t.ClientID = c.ClientID
            WHERE t.ManagerID = ?
        """
        params = [manager_id]
        if client_id_filter:
            query += " AND t.ClientID = ?"
            params.append(client_id_filter)
        if only_pending == '1':
            query += " AND t.IsCompleted = 0"
        query += " ORDER BY t.ScheduledAt ASC"
        cur.execute(query, *params)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>/tasks', methods=['POST'])
@login_required
def create_task(client_id):
    data = request.get_json()
    if not data.get('task_type') or not data.get('title') or not data.get('scheduled_at'):
        return jsonify({'error': 'Тип, заголовок и время задачи обязательны'}), 400

    allowed_types = ('call', 'meeting', 'letter')
    if data['task_type'] not in allowed_types:
        return jsonify({'error': f'Тип задачи должен быть одним из: {", ".join(allowed_types)}'}), 400

    try:
        # Конвертируем строку даты/времени в объект datetime для pyodbc
        raw_sched = data['scheduled_at']
        try:
            raw_sched = raw_sched.replace('T', ' ')
            if len(raw_sched) == 16:
                raw_sched += ':00'
            scheduled_dt = datetime.strptime(raw_sched, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            scheduled_dt = raw_sched

        # Проверяем что дата не в прошлом (только дата, не время)
        if isinstance(scheduled_dt, datetime):
            if scheduled_dt.date() < date.today():
                return jsonify({'error': 'Нельзя планировать задачу на прошедшую дату'}), 400

        conn = get_connection()
        cur = conn.cursor()
        deal_id = data.get('deal_id') or None
        cur.execute("""
            INSERT INTO Tasks (ClientID, ManagerID, DealID, TaskType, Title, Description, ScheduledAt)
            OUTPUT INSERTED.TaskID, INSERTED.CreatedAt
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, client_id, session['user_id'], deal_id,
             data['task_type'], data['title'],
             data.get('description'), scheduled_dt)
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'task_id': row[0], 'created_at': row[1].isoformat()}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks/<int:task_id>/complete', methods=['POST'])
@login_required
def complete_task(task_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT t.ManagerID, t.ClientID, t.DealID, t.TaskType, t.Title
            FROM Tasks t WHERE t.TaskID=?
        """, task_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Задача не найдена'}), 404
        if session['role'] != 'admin' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        manager_id = row[0]
        client_id  = row[1]
        deal_id    = row[2]
        task_type  = row[3]
        task_title = row[4]

        type_labels = {'call': 'Звонок', 'meeting': 'Встреча', 'letter': 'Письмо'}
        type_label = type_labels.get(task_type, task_type)

        cur.execute("UPDATE Tasks SET IsCompleted=1, CompletedAt=GETDATE() WHERE TaskID=?", task_id)

        now = datetime.now()

        # Добавляем событие в ленту событий клиента (IsAutomatic=1, нельзя удалить менеджеру)
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
            VALUES (?, ?, 'task_completed', ?, ?, ?, 1, ?)
        """, client_id, session['user_id'],
             now,
             f'Выполнена задача: {task_title}',
             f'Тип: {type_label}',
             deal_id)

        # Если задача из сделки — добавляем событие и в ленту сделки
        if deal_id:
            cur.execute("""
                INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
                VALUES (?, 'task_completed', ?, 1, ?)
            """, deal_id, f'Выполнена задача: {task_title} ({type_label})', now)

        conn.commit()
        conn.close()
        return jsonify({'message': 'Задача выполнена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>/all-tasks', methods=['GET'])
@login_required
def get_client_all_tasks(client_id):
    """Все задачи клиента: собственные + задачи из сделок клиента."""
    only_pending = request.args.get('pending')
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Задачи клиента (таблица Tasks)
        q1 = """
            SELECT t.TaskID, t.TaskType, t.Title, t.Description,
                   t.ScheduledAt, t.CompletedAt, t.IsCompleted, t.CreatedAt,
                   t.DealID, 'client' AS TaskSource, NULL AS DealTitle
            FROM Tasks t
            WHERE t.ClientID = ?
        """
        params1 = [client_id]
        if only_pending == '1':
            q1 += " AND t.IsCompleted = 0"
        q1 += " ORDER BY t.ScheduledAt ASC"
        cur.execute(q1, *params1)
        client_tasks = rows_to_list(cur, cur.fetchall())

        # Задачи из сделок клиента (таблица DealTasks)
        q2 = """
            SELECT dt.TaskID, dt.TaskType, dt.Title, dt.Description,
                   dt.ScheduledAt, dt.CompletedAt, dt.IsCompleted, dt.CreatedAt,
                   dt.DealID, 'deal' AS TaskSource, d.Title AS DealTitle
            FROM DealTasks dt
            INNER JOIN Deals d ON dt.DealID = d.DealID
            WHERE d.ClientID = ?
        """
        params2 = [client_id]
        if only_pending == '1':
            q2 += " AND dt.IsCompleted = 0"
        q2 += " ORDER BY dt.ScheduledAt ASC"
        cur.execute(q2, *params2)
        deal_tasks = rows_to_list(cur, cur.fetchall())

        conn.close()
        return jsonify({'client_tasks': client_tasks, 'deal_tasks': deal_tasks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT ManagerID FROM Tasks WHERE TaskID=?", task_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Задача не найдена'}), 404
        if session['role'] != 'admin' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM Tasks WHERE TaskID=?", task_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Задача удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Оповещения ──────────────────────────────────────────────────────────────

@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    manager_id = session['user_id']
    today = date.today()
    try:
        conn = get_connection()
        cur = conn.cursor()

        notifications = []
        type_labels = {'call': 'Звонок', 'meeting': 'Встреча', 'letter': 'Письмо'}

        # Просроченные задачи клиента (красный)
        cur.execute("""
            SELECT t.TaskID, t.TaskType, t.Title, t.ScheduledAt,
                   c.FullName AS ClientName, c.ClientID, t.DealID
            FROM Tasks t
            INNER JOIN Clients c ON t.ClientID = c.ClientID
            WHERE t.ManagerID = ? AND t.IsCompleted = 0
              AND CAST(t.ScheduledAt AS DATE) < ?
            ORDER BY t.ScheduledAt ASC
        """, manager_id, today.isoformat())
        for row in cur.fetchall():
            r = row_to_dict(cur, row)
            notifications.append({
                'type': 'overdue',
                'color': 'red',
                'task_id': r['TaskID'],
                'task_source': 'client',
                'task_type': r['TaskType'],
                'task_type_label': type_labels.get(r['TaskType'], r['TaskType']),
                'title': r['Title'],
                'scheduled_at': r['ScheduledAt'],
                'client_name': r['ClientName'],
                'client_id': r['ClientID'],
                'deal_id': r['DealID'],
            })

        # Просроченные задачи по сделкам (красный)
        cur.execute("""
            SELECT dt.TaskID, dt.TaskType, dt.Title, dt.ScheduledAt,
                   c.FullName AS ClientName, c.ClientID, d.DealID, d.Title AS DealTitle
            FROM DealTasks dt
            INNER JOIN Deals d ON dt.DealID = d.DealID
            INNER JOIN Clients c ON d.ClientID = c.ClientID
            WHERE d.ManagerID = ? AND dt.IsCompleted = 0
              AND CAST(dt.ScheduledAt AS DATE) < ?
            ORDER BY dt.ScheduledAt ASC
        """, manager_id, today.isoformat())
        for row in cur.fetchall():
            r = row_to_dict(cur, row)
            notifications.append({
                'type': 'overdue',
                'color': 'red',
                'task_id': r['TaskID'],
                'task_source': 'deal',
                'task_type': r['TaskType'],
                'task_type_label': type_labels.get(r['TaskType'], r['TaskType']),
                'title': r['Title'],
                'scheduled_at': r['ScheduledAt'],
                'client_name': r['ClientName'],
                'client_id': r['ClientID'],
                'deal_id': r['DealID'],
                'deal_title': r['DealTitle'],
            })

        # Задачи клиента на сегодня (жёлтый)
        cur.execute("""
            SELECT t.TaskID, t.TaskType, t.Title, t.ScheduledAt,
                   c.FullName AS ClientName, c.ClientID, t.DealID
            FROM Tasks t
            INNER JOIN Clients c ON t.ClientID = c.ClientID
            WHERE t.ManagerID = ? AND t.IsCompleted = 0
              AND CAST(t.ScheduledAt AS DATE) = ?
            ORDER BY t.ScheduledAt ASC
        """, manager_id, today.isoformat())
        for row in cur.fetchall():
            r = row_to_dict(cur, row)
            notifications.append({
                'type': 'today',
                'color': 'yellow',
                'task_id': r['TaskID'],
                'task_source': 'client',
                'task_type': r['TaskType'],
                'task_type_label': type_labels.get(r['TaskType'], r['TaskType']),
                'title': r['Title'],
                'scheduled_at': r['ScheduledAt'],
                'client_name': r['ClientName'],
                'client_id': r['ClientID'],
                'deal_id': r['DealID'],
            })

        # Задачи по сделкам на сегодня (жёлтый)
        cur.execute("""
            SELECT dt.TaskID, dt.TaskType, dt.Title, dt.ScheduledAt,
                   c.FullName AS ClientName, c.ClientID, d.DealID, d.Title AS DealTitle
            FROM DealTasks dt
            INNER JOIN Deals d ON dt.DealID = d.DealID
            INNER JOIN Clients c ON d.ClientID = c.ClientID
            WHERE d.ManagerID = ? AND dt.IsCompleted = 0
              AND CAST(dt.ScheduledAt AS DATE) = ?
            ORDER BY dt.ScheduledAt ASC
        """, manager_id, today.isoformat())
        for row in cur.fetchall():
            r = row_to_dict(cur, row)
            notifications.append({
                'type': 'today',
                'color': 'yellow',
                'task_id': r['TaskID'],
                'task_source': 'deal',
                'task_type': r['TaskType'],
                'task_type_label': type_labels.get(r['TaskType'], r['TaskType']),
                'title': r['Title'],
                'scheduled_at': r['ScheduledAt'],
                'client_name': r['ClientName'],
                'client_id': r['ClientID'],
                'deal_id': r['DealID'],
                'deal_title': r['DealTitle'],
            })

        # Дни рождения сегодня и в ближайшие 7 дней
        cur.execute("""
            SELECT c.ClientID, c.FullName, c.BirthDate, c.Phone
            FROM Clients c
            WHERE c.ManagerID = ?
              AND c.BirthDate IS NOT NULL
              AND (
                  (MONTH(c.BirthDate) = MONTH(GETDATE()) AND DAY(c.BirthDate) = DAY(GETDATE()))
                  OR
                  (DATEADD(YEAR,
                    DATEDIFF(YEAR, c.BirthDate, GETDATE())
                      + CASE WHEN DATEFROMPARTS(YEAR(GETDATE()), MONTH(c.BirthDate), DAY(c.BirthDate)) < CAST(GETDATE() AS DATE) THEN 1 ELSE 0 END,
                    c.BirthDate) BETWEEN CAST(GETDATE() AS DATE) AND DATEADD(DAY, 7, CAST(GETDATE() AS DATE))
                  )
              )
        """, manager_id)
        for row in cur.fetchall():
            r = row_to_dict(cur, row)
            bday_this_year = None
            if r.get('BirthDate'):
                try:
                    bd = date.fromisoformat(r['BirthDate'][:10])
                    bday_this_year = date(today.year, bd.month, bd.day)
                    if bday_this_year < today:
                        bday_this_year = date(today.year + 1, bd.month, bd.day)
                except Exception:
                    pass
            is_today = bday_this_year == today if bday_this_year else False
            notifications.append({
                'type': 'birthday',
                'color': 'blue',
                'client_id': r['ClientID'],
                'client_name': r['FullName'],
                'birth_date': r['BirthDate'],
                'birthday_this_year': bday_this_year.isoformat() if bday_this_year else None,
                'is_today': is_today,
                'title': f'День рождения: {r["FullName"]}' + (' — сегодня!' if is_today else ''),
            })

        conn.close()
        return jsonify(notifications)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# DEALS (Сделки)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/deals', methods=['GET'])
@login_required
def get_deals():
    manager_id_filter = request.args.get('manager_id')
    client_id_filter  = request.args.get('client_id')
    stage_id_filter   = request.args.get('stage_id')
    archived_filter   = request.args.get('archived', '0')

    try:
        conn = get_connection()
        cur = conn.cursor()

        query = """
            SELECT d.DealID, d.Title, d.DealType, d.Budget, d.Priority,
                   d.DealDate, d.Deadline, d.Notes AS DealNotes, d.IsArchived,
                   d.CreatedAt, d.UpdatedAt, d.CompletedAt,
                   ds.StageID, ds.StageName, ds.StageOrder, ds.IsCompleted,
                   c.ClientID, c.FullName AS ClientName, c.Phone AS ClientPhone,
                   c.CompanyName AS ClientCompany,
                   u.UserID AS ManagerID, u.FullName AS ManagerName
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            INNER JOIN Clients c ON d.ClientID = c.ClientID
            INNER JOIN Users u ON d.ManagerID = u.UserID
            WHERE 1=1
        """
        params = []

        if session['role'] == 'manager':
            query += " AND d.ManagerID = ?"
            params.append(session['user_id'])
        elif manager_id_filter:
            query += " AND d.ManagerID = ?"
            params.append(manager_id_filter)

        if client_id_filter:
            query += " AND d.ClientID = ?"
            params.append(client_id_filter)

        if stage_id_filter:
            query += " AND d.StageID = ?"
            params.append(stage_id_filter)

        query += " AND d.IsArchived = ?"
        params.append(1 if archived_filter == '1' else 0)

        query += " ORDER BY d.UpdatedAt DESC"

        cur.execute(query, *params)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals', methods=['POST'])
@manager_required
def create_deal():
    data = request.get_json()
    required = ['client_id', 'title']
    for f in required:
        if not data.get(f):
            return jsonify({'error': f'Поле {f} обязательно'}), 400

    manager_id = session['user_id']
    try:
        conn = get_connection()
        cur = conn.cursor()

        if session['role'] == 'manager':
            cur.execute("SELECT ManagerID FROM Clients WHERE ClientID=?", data['client_id'])
            row = cur.fetchone()
            if not row or row[0] != manager_id:
                conn.close()
                return jsonify({'error': 'Нет доступа к этому клиенту'}), 403

        # Новые сделки создаются на этапе 1 ("Новый контакт")
        # Поля типа, приоритета, бюджета заполняются позже
        cur.execute("""
            SET NOCOUNT ON;
            INSERT INTO Deals (ClientID, ManagerID, StageID, DealType, Title, Budget, Priority, DealDate, Notes)
            VALUES (?, ?, 1, 'розница', ?, NULL, 2, ?, ?);
            SELECT CAST(SCOPE_IDENTITY() AS INT);
        """,
            data['client_id'], manager_id,
            data['title'],
            data.get('deal_date', date.today().isoformat()),
            data.get('notes')
        )
        row = cur.fetchone()
        deal_id = int(row[0])

        now = datetime.now()
        client_id = data['client_id']

        # Записываем первое автоматическое событие в ленту сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'created', ?, 1, ?)
        """, deal_id, f'Сделка создана: {data["title"]}', now)

        # Дублируем событие создания сделки в ленту клиента
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
            VALUES (?, ?, 'deal_stage', ?, ?, ?, 1, ?)
        """, client_id, manager_id, now,
             'Создана новая сделка',
             f'Сделка "{data["title"]}" создана и добавлена на этап "Новый контакт"',
             deal_id)

        conn.commit()
        conn.close()
        return jsonify({'deal_id': deal_id, 'message': 'Сделка создана'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>', methods=['GET'])
@login_required
def get_deal(deal_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT d.DealID, d.Title, d.DealType, d.Budget, d.Priority,
                   d.DealDate, d.Deadline, d.Notes AS DealNotes, d.IsArchived,
                   d.CreatedAt, d.UpdatedAt,
                   ds.StageID, ds.StageName, ds.StageOrder, ds.IsCompleted,
                   c.ClientID, c.FullName AS ClientName,
                   u.UserID AS ManagerID, u.FullName AS ManagerName
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            INNER JOIN Clients c ON d.ClientID = c.ClientID
            INNER JOIN Users u ON d.ManagerID = u.UserID
            WHERE d.DealID = ?
        """, deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        deal = row_to_dict(cur, row)
        conn.close()
        return jsonify(deal)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>', methods=['PUT'])
@login_required
def update_deal(deal_id):
    data = request.get_json()
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT ManagerID FROM Deals WHERE DealID=?", deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        cur.execute("""
            UPDATE Deals SET
                DealType  = ISNULL(?, DealType),
                Title     = ISNULL(?, Title),
                Budget    = ISNULL(?, Budget),
                Priority  = ISNULL(?, Priority),
                DealDate  = ISNULL(?, DealDate),
                Notes     = ?,
                UpdatedAt = GETDATE()
            WHERE DealID = ?
        """,
            data.get('deal_type'),
            data.get('title'),
            data.get('budget') or None,
            data.get('priority'),
            data.get('deal_date'),
            data.get('notes'),
            deal_id
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Сделка обновлена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>/stage', methods=['POST'])
@login_required
def update_deal_stage(deal_id):
    data = request.get_json()
    stage_id = data.get('stage_id')
    stage_deadline = data.get('stage_deadline')  # дата конечного срока выполнения этапа
    if not stage_id:
        return jsonify({'error': 'Укажите этап'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Получаем текущие данные сделки
        cur.execute("""
            SELECT d.ManagerID, d.StageID, d.ClientID, d.Title, d.DealType, d.Budget,
                   ds.StageOrder AS CurrentStageOrder, ds.StageName AS CurrentStageName,
                   d.DealID
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.DealID=?
        """, deal_id)
        deal_row = cur.fetchone()
        if not deal_row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404

        manager_id_deal = deal_row[0]
        current_stage_id = deal_row[1]
        client_id = deal_row[2]
        deal_title = deal_row[3]
        current_stage_order = deal_row[6]
        current_stage_name = deal_row[7]

        if session['role'] == 'manager' and manager_id_deal != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        # Получаем данные нового этапа
        cur.execute("SELECT StageName, StageOrder, IsCompleted FROM DealStages WHERE StageID=?", stage_id)
        new_stage_row = cur.fetchone()
        if not new_stage_row:
            conn.close()
            return jsonify({'error': 'Этап не найден'}), 404

        new_stage_name = new_stage_row[0]
        new_stage_order = new_stage_row[1]
        new_is_completed = bool(new_stage_row[2])

        # Менеджер не может переносить сделку назад
        if session['role'] == 'manager' and new_stage_order < current_stage_order:
            conn.close()
            return jsonify({'error': 'Нельзя перенести сделку на предыдущий этап'}), 400

        # Нельзя перенести более чем на 1 этап вперёд
        if new_stage_order > current_stage_order + 1:
            conn.close()
            return jsonify({'error': 'Нельзя перенести сделку более чем на 1 этап вперёд'}), 400

        # Если тот же этап — ничего не делаем
        if stage_id == current_stage_id:
            conn.close()
            return jsonify({'message': 'Этап не изменился'})

        # === Валидация условий перехода ===
        # Переход с "Выявление потребностей" (order=2) на "Подбор и расчёт" (order=3)
        # Требует: DealType, Priority, Budget
        if current_stage_order == 2 and new_stage_order == 3:
            cur.execute("SELECT DealType, Priority, Budget FROM Deals WHERE DealID=?", deal_id)
            r = cur.fetchone()
            if not r or not r[0] or r[1] is None:
                conn.close()
                return jsonify({'error': 'Перед переносом заполните поля: тип сделки, приоритет и бюджет в карточке сделки'}), 400
            if not r[2]:
                conn.close()
                return jsonify({'error': 'Перед переносом укажите бюджет в карточке сделки'}), 400

        # Переход с "Подбор и расчёт" (order=3) и далее требует товаров в составе заказа
        if current_stage_order >= 3 and new_stage_order > current_stage_order:
            cur.execute("SELECT COUNT(*) FROM DealItems WHERE DealID=?", deal_id)
            items_count = cur.fetchone()[0]
            if items_count == 0:
                conn.close()
                return jsonify({'error': 'Необходимо добавить хотя бы одну позицию в состав заказа'}), 400

        # Проверка что дата передана для этапов 2→3, 3→4, 4→5, 5→6, 6→7 (не для 7→8)
        if current_stage_order >= 2 and new_stage_order <= 7 and not new_is_completed:
            if not stage_deadline:
                conn.close()
                return jsonify({'error': 'Укажите конечный срок выполнения этапа'}), 400

        # Обновляем этап
        new_deadline = stage_deadline if stage_deadline else None
        if new_is_completed:
            cur.execute("""
                UPDATE Deals
                SET StageID=?, UpdatedAt=GETDATE(), CompletedAt=GETDATE(), Deadline=?
                WHERE DealID=?
            """, stage_id, new_deadline, deal_id)
        else:
            cur.execute("""
                UPDATE Deals
                SET StageID=?, UpdatedAt=GETDATE(), CompletedAt=NULL, Deadline=?
                WHERE DealID=?
            """, stage_id, new_deadline, deal_id)

        now = datetime.now()

        # Записываем автоматическое событие в ленту сделки
        event_desc = f'Этап изменён: {current_stage_name} → {new_stage_name}'
        if stage_deadline:
            event_desc += f'. Срок выполнения этапа: {stage_deadline}'
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'stage_change', ?, 1, ?)
        """, deal_id, event_desc, now)

        # Если сделка завершена — добавляем событие "deal_closed" в ленте сделки и клиента
        if new_is_completed:
            cur.execute("""
                INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
                VALUES (?, 'closed', ?, 1, ?)
            """, deal_id, f'Сделка завершена: {deal_title}', now)

            # В ленту клиента
            cur.execute("""
                INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
                VALUES (?, ?, 'deal_closed', ?, ?, ?, 1, ?)
            """, client_id, session['user_id'], now,
                 f'Сделка завершена', f'Сделка "{deal_title}" переведена в статус "Выполнено"', deal_id)
        else:
            # Дублируем событие смены этапа в ленту клиента
            cur.execute("""
                INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
                VALUES (?, ?, 'deal_stage', ?, ?, ?, 1, ?)
            """, client_id, session['user_id'], now,
                 f'Изменён этап сделки', event_desc, deal_id)

        # Если переход на "Счёт выставлен" (order=5) — генерируем счёт автоматически
        if new_stage_order == 5:
            try:
                _generate_invoice(conn, cur, deal_id, client_id)
            except Exception:
                pass  # Не критично если не удалось сгенерировать счёт

        conn.commit()
        conn.close()
        return jsonify({'message': 'Этап обновлён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _generate_invoice(conn, cur, deal_id, client_id):
    """Генерирует счёт на оплату в формате HTML и сохраняет как документ."""
    try:

        # Получаем данные сделки
        cur.execute("""
            SELECT d.DealID, d.Title, d.DealDate,
                   c.FullName AS ClientName, c.CompanyName, c.Phone, c.Email, c.Address,
                   u.FullName AS ManagerName
            FROM Deals d
            INNER JOIN Clients c ON d.ClientID = c.ClientID
            INNER JOIN Users u ON d.ManagerID = u.UserID
            WHERE d.DealID = ?
        """, deal_id)
        deal_row = cur.fetchone()
        if not deal_row:
            return

        # Получаем позиции
        cur.execute("""
            SELECT p.ProductName, p.SKU, p.Unit, di.Quantity, di.UnitPrice, di.LineTotal,
                   ISNULL(p.TileWidth,0), ISNULL(p.TileLength,0)
            FROM DealItems di
            INNER JOIN Products p ON di.ProductID = p.ProductID
            WHERE di.DealID = ?
            ORDER BY di.DealItemID
        """, deal_id)
        items = cur.fetchall()

        total = sum(float(i[5] or 0) for i in items)
        invoice_date = date.today().strftime('%d.%m.%Y')

        # Строки позиций
        items_rows = ''
        for idx, i in enumerate(items, 1):
            qty = float(i[3])
            tw = float(i[6])
            tl = float(i[7])
            vol = qty * tw * tl / 1000000.0 if tw and tl else 0.0
            items_rows += f'''
            <tr>
                <td style="border:1px solid #ccc;padding:6px;text-align:center">{idx}</td>
                <td style="border:1px solid #ccc;padding:6px">{i[0]}</td>
                <td style="border:1px solid #ccc;padding:6px;text-align:center">{i[2]}</td>
                <td style="border:1px solid #ccc;padding:6px;text-align:right">{qty:.0f}</td>
                <td style="border:1px solid #ccc;padding:6px;text-align:right">{vol:.4f}</td>
                <td style="border:1px solid #ccc;padding:6px;text-align:right">{float(i[4]):.2f} BYN</td>
                <td style="border:1px solid #ccc;padding:6px;text-align:right;font-weight:bold">{float(i[5]):.2f} BYN</td>
            </tr>'''

        total_vol = sum(float(i[3]) * float(i[6]) * float(i[7]) / 1000000.0 for i in items if float(i[6]) and float(i[7]))

        html_content = f"""<!DOCTYPE html>
<html lang="ru">
<head><meta charset="UTF-8"><title>Счёт на оплату #{deal_id}</title>
<style>body{{font-family:Arial,sans-serif;font-size:12px;margin:20px}}</style>
</head>
<body>
<div style="max-width:900px;margin:0 auto">
  <div style="display:flex;justify-content:space-between;margin-bottom:20px">
    <div>
      <h2 style="margin:0;color:#1e40af">ООО "Фэшн Галлери"</h2>
      <p style="margin:4px 0;color:#64748b">220040 г. Минск, ул. Богдановича 108, пом. 22</p>
      <p style="margin:4px 0;color:#64748b">Тел: +375(29)1111176</p>
      <p style="margin:4px 0;color:#64748b">УНП: 192697945</p>
    </div>
    <div style="text-align:right">
      <h1 style="margin:0;font-size:1.5rem">СЧЁТ НА ОПЛАТУ</h1>
      <p style="margin:4px 0">№ {deal_id} от {invoice_date}</p>
    </div>
  </div>
  <hr style="border:2px solid #1e40af;margin-bottom:16px">
  <table style="width:100%;margin-bottom:16px">
    <tr>
      <td style="width:50%;vertical-align:top">
        <strong>Поставщик:</strong><br>
        ООО "Фэшн Галлери"<br>
        УНП 192697945<br>
        220040 г. Минск, ул. Богдановича 108, пом. 22<br>
        Тел: +375(29)1111176
      </td>
      <td style="width:50%;vertical-align:top">
        <strong>Покупатель:</strong><br>
        {deal_row[3]}{(' (' + deal_row[4] + ')') if deal_row[4] else ''}<br>
        {deal_row[7] or ''}<br>
        {deal_row[6] or ''}<br>
        {deal_row[5] or ''}
      </td>
    </tr>
  </table>
  <p><strong>Основание:</strong> Сделка №{deal_id}: {deal_row[1]}</p>
  <table style="width:100%;border-collapse:collapse;margin-bottom:16px">
    <thead>
      <tr style="background:#f1f5f9">
        <th style="border:1px solid #ccc;padding:6px;text-align:center">№</th>
        <th style="border:1px solid #ccc;padding:6px;text-align:left">Наименование</th>
        <th style="border:1px solid #ccc;padding:6px;text-align:center">Ед.изм.</th>
        <th style="border:1px solid #ccc;padding:6px;text-align:right">Кол-во (шт.)</th>
        <th style="border:1px solid #ccc;padding:6px;text-align:right">Объём (м²)</th>
        <th style="border:1px solid #ccc;padding:6px;text-align:right">Цена за шт. (BYN)</th>
        <th style="border:1px solid #ccc;padding:6px;text-align:right">Сумма (BYN)</th>
      </tr>
    </thead>
    <tbody>
      {items_rows}
    </tbody>
    <tfoot>
      <tr>
        <td colspan="4" style="border:1px solid #ccc;padding:6px;text-align:right;font-weight:bold">ИТОГО:</td>
        <td style="border:1px solid #ccc;padding:6px;text-align:right;font-weight:bold">{total_vol:.4f} м²</td>
        <td></td>
        <td style="border:1px solid #ccc;padding:6px;text-align:right;font-weight:bold;color:#16a34a">{total:.2f} BYN</td>
      </tr>
    </tfoot>
  </table>
  <p style="color:#64748b;font-size:11px">НДС не предусмотрен.</p>
  <div style="margin-top:40px;display:flex;justify-content:space-between">
    <div>Руководитель: _________________ / ООО "Фэшн Галлери" /</div>
    <div>Бухгалтер: _________________</div>
  </div>
</div>
</body>
</html>"""

        # Сохраняем как файл
        stored_name = f"invoice_{deal_id}_{uuid.uuid4().hex[:8]}.html"
        deal_upload_dir = os.path.join(UPLOAD_DIR, 'deals', str(deal_id))
        os.makedirs(deal_upload_dir, exist_ok=True)
        file_path = os.path.join(deal_upload_dir, stored_name)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        file_size = os.path.getsize(file_path)
        original_name = f'Счёт на оплату №{deal_id} от {invoice_date}.html'

        # Сохраняем в DealDocuments
        cur.execute("""
            INSERT INTO DealDocuments (DealID, UploadedBy, FileName, OriginalName, FileSize, MimeType)
            OUTPUT INSERTED.DocID
            VALUES (?, 1, ?, ?, ?, 'text/html')
        """, deal_id, stored_name, original_name, file_size)
        doc_row = cur.fetchone()
        doc_id = doc_row[0] if doc_row else None

        now = datetime.now()

        # Событие в ленте сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'doc_added', ?, 1, ?)
        """, deal_id, f'Добавлен документ: {original_name}', now)

        # Событие в ленте клиента
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
            VALUES (?, 1, 'doc_added', ?, ?, ?, 1, ?)
        """, client_id, now,
             f'Добавлен документ',
             f'В сделке автоматически создан счёт на оплату: {original_name}',
             deal_id)

        # NOTE: commit будет сделан в вызывающей функции
    except Exception as e:
        pass  # Не критично


@app.route('/api/deals/<int:deal_id>', methods=['DELETE'])
@login_required
def delete_deal(deal_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT ManagerID FROM Deals WHERE DealID=?", deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM Deals WHERE DealID=?", deal_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Сделка удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deal-stages', methods=['GET'])
@login_required
def get_deal_stages():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT StageID, StageName, StageOrder, IsCompleted FROM DealStages ORDER BY StageOrder")
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Полная карточка сделки ───────────────────────────────────────────────────

@app.route('/api/deals/<int:deal_id>/full', methods=['GET'])
@login_required
def get_deal_full(deal_id):
    """Полная карточка сделки: инфо + позиции + события + задачи + документы."""
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Основная информация
        cur.execute("""
            SELECT d.DealID, d.Title, d.DealType, d.Budget, d.Priority,
                   d.DealDate, d.Deadline, d.Notes AS DealNotes, d.IsArchived,
                   d.CreatedAt, d.UpdatedAt, d.CompletedAt,
                   ds.StageID, ds.StageName, ds.StageOrder, ds.IsCompleted,
                   c.ClientID, c.FullName AS ClientName, c.Phone AS ClientPhone,
                   c.CompanyName AS ClientCompany,
                   u.UserID AS ManagerID, u.FullName AS ManagerName,
                   d.AppliedOfferID
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            INNER JOIN Clients c ON d.ClientID = c.ClientID
            INNER JOIN Users u ON d.ManagerID = u.UserID
            WHERE d.DealID = ?
        """, deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404

        deal = row_to_dict(cur, row)

        if session['role'] == 'manager' and deal['ManagerID'] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        # Позиции заказа
        cur.execute("""
            SELECT di.DealItemID, di.ProductID, p.ProductName, p.SKU, p.Unit,
                   di.Quantity, di.UnitPrice, di.LineTotal,
                   p.TileWidth, p.TileLength, p.TileThickness
            FROM DealItems di
            INNER JOIN Products p ON di.ProductID = p.ProductID
            WHERE di.DealID = ?
            ORDER BY di.DealItemID
        """, deal_id)
        items = rows_to_list(cur, cur.fetchall())

        # События (лента сделки)
        cur.execute("""
            SELECT de.EventID, de.EventType, de.EventDate, de.Description, de.IsAutomatic
            FROM DealEvents de
            WHERE de.DealID = ?
            ORDER BY de.EventDate DESC
        """, deal_id)
        events = rows_to_list(cur, cur.fetchall())

        # Задачи по сделке
        cur.execute("""
            SELECT dt.TaskID, dt.TaskType, dt.Title, dt.Description,
                   dt.ScheduledAt, dt.CompletedAt, dt.IsCompleted, dt.CreatedAt
            FROM DealTasks dt
            WHERE dt.DealID = ?
            ORDER BY dt.ScheduledAt ASC
        """, deal_id)
        tasks = rows_to_list(cur, cur.fetchall())

        # Документы
        cur.execute("""
            SELECT dd.DocID, dd.FileName, dd.OriginalName, dd.FileSize, dd.MimeType,
                   dd.UploadedAt, u.FullName AS UploaderName
            FROM DealDocuments dd
            INNER JOIN Users u ON dd.UploadedBy = u.UserID
            WHERE dd.DealID = ?
            ORDER BY dd.UploadedAt DESC
        """, deal_id)
        documents = rows_to_list(cur, cur.fetchall())

        # Applied offer details
        applied_offer = None
        if deal.get('AppliedOfferID'):
            cur.execute("""
                SELECT co.OfferID, co.OfferNumber, co.Status, co.ExpiresAt, co.Notes,
                       u2.FullName AS CreatedByName
                FROM CommercialOffers co
                INNER JOIN Users u2 ON co.CreatedBy = u2.UserID
                WHERE co.OfferID = ?
            """, deal['AppliedOfferID'])
            offer_row = cur.fetchone()
            if offer_row:
                applied_offer = row_to_dict(cur, offer_row)
                # Get offer items
                cur.execute("""
                    SELECT coi.ItemID, coi.ProductID, p.ProductName, p.SKU,
                           p.Price AS BasePrice, coi.DiscountPct,
                           ROUND(p.Price * (1 - coi.DiscountPct/100.0), 2) AS DiscountedPrice,
                           p.TileWidth, p.TileLength
                    FROM CommercialOfferItems coi
                    INNER JOIN Products p ON coi.ProductID = p.ProductID
                    WHERE coi.OfferID = ?
                """, deal['AppliedOfferID'])
                applied_offer['items'] = rows_to_list(cur, cur.fetchall())

        conn.close()
        deal['items']          = items
        deal['events']         = events
        deal['tasks']          = tasks
        deal['documents']      = documents
        deal['applied_offer']  = applied_offer
        return jsonify(deal)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Позиции сделки (DealItems) ───────────────────────────────────────────────

@app.route('/api/deals/<int:deal_id>/items', methods=['GET'])
@login_required
def get_deal_items(deal_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT di.DealItemID, di.ProductID, p.ProductName, p.SKU, p.Unit,
                   di.Quantity, di.UnitPrice, di.LineTotal,
                   p.TileWidth, p.TileLength, p.TileThickness
            FROM DealItems di
            INNER JOIN Products p ON di.ProductID = p.ProductID
            WHERE di.DealID = ?
            ORDER BY di.DealItemID
        """, deal_id)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>/items', methods=['POST'])
@login_required
def add_deal_item(deal_id):
    data = request.get_json()
    if not data.get('product_id') or not data.get('quantity') or not data.get('unit_price'):
        return jsonify({'error': 'Обязательны: product_id, quantity, unit_price'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT ManagerID FROM Deals WHERE DealID=?", deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        # Проверяем, не добавлен ли уже этот товар
        cur.execute("SELECT DealItemID FROM DealItems WHERE DealID=? AND ProductID=?",
                    deal_id, data['product_id'])
        existing = cur.fetchone()
        if existing:
            # Обновляем существующую позицию
            cur.execute("""
                UPDATE DealItems SET Quantity=?, UnitPrice=?
                WHERE DealID=? AND ProductID=?
            """, float(data['quantity']), float(data['unit_price']), deal_id, data['product_id'])
        else:
            cur.execute("""
                INSERT INTO DealItems (DealID, ProductID, Quantity, UnitPrice)
                VALUES (?, ?, ?, ?)
            """, deal_id, data['product_id'], float(data['quantity']), float(data['unit_price']))

        # Пересчитываем бюджет сделки
        cur.execute("""
            UPDATE Deals SET
                Budget = (SELECT ISNULL(SUM(LineTotal), 0) FROM DealItems WHERE DealID = ?),
                UpdatedAt = GETDATE()
            WHERE DealID = ?
        """, deal_id, deal_id)

        # Записываем событие
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description)
            VALUES (?, 'item_added', ?)
        """, deal_id, f"Добавлена позиция: {data.get('product_name', '')} × {data['quantity']}")

        conn.commit()
        conn.close()
        return jsonify({'message': 'Позиция добавлена'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deal-items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_deal_item(item_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT di.DealID, d.ManagerID FROM DealItems di INNER JOIN Deals d ON di.DealID=d.DealID WHERE di.DealItemID=?", item_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Позиция не найдена'}), 404
        deal_id = row[0]
        if session['role'] == 'manager' and row[1] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        cur.execute("DELETE FROM DealItems WHERE DealItemID=?", item_id)
        # Пересчёт бюджета
        cur.execute("""
            UPDATE Deals SET
                Budget = (SELECT ISNULL(SUM(LineTotal), 0) FROM DealItems WHERE DealID = ?),
                UpdatedAt = GETDATE()
            WHERE DealID = ?
        """, deal_id, deal_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Позиция удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Задачи по сделке (DealTasks) ────────────────────────────────────────────

@app.route('/api/deals/<int:deal_id>/tasks', methods=['GET'])
@login_required
def get_deal_tasks(deal_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT TaskID, TaskType, Title, Description, ScheduledAt, CompletedAt, IsCompleted, CreatedAt
            FROM DealTasks
            WHERE DealID = ?
            ORDER BY ScheduledAt ASC
        """, deal_id)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>/tasks', methods=['POST'])
@login_required
def create_deal_task(deal_id):
    data = request.get_json()
    if not data.get('task_type') or not data.get('title') or not data.get('scheduled_at'):
        return jsonify({'error': 'Тип, заголовок и время обязательны'}), 400

    allowed_types = ('call', 'meeting', 'letter')
    if data['task_type'] not in allowed_types:
        return jsonify({'error': f'Тип задачи должен быть одним из: {", ".join(allowed_types)}'}), 400

    try:
        raw_sched = data['scheduled_at'].replace('T', ' ')
        if len(raw_sched) == 16:
            raw_sched += ':00'
        try:
            scheduled_dt = datetime.strptime(raw_sched, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            scheduled_dt = raw_sched

        # Проверяем что дата не в прошлом
        if isinstance(scheduled_dt, datetime):
            if scheduled_dt.date() < date.today():
                return jsonify({'error': 'Нельзя планировать задачу на прошедшую дату'}), 400

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT ManagerID, ClientID FROM Deals WHERE DealID=?", deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        client_id = row[1]

        cur.execute("""
            INSERT INTO DealTasks (DealID, TaskType, Title, Description, ScheduledAt)
            OUTPUT INSERTED.TaskID, INSERTED.CreatedAt
            VALUES (?, ?, ?, ?, ?)
        """, deal_id, data['task_type'], data['title'], data.get('description'), scheduled_dt)
        row = cur.fetchone()
        new_task_id = row[0]

        conn.commit()
        conn.close()
        return jsonify({'task_id': new_task_id, 'created_at': row[1].isoformat()}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deal-tasks/<int:task_id>/complete', methods=['POST'])
@login_required
def complete_deal_task(task_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT dt.DealID, d.ManagerID, d.ClientID, dt.TaskType, dt.Title
            FROM DealTasks dt
            INNER JOIN Deals d ON dt.DealID = d.DealID
            WHERE dt.TaskID=?
        """, task_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Задача не найдена'}), 404
        if session['role'] == 'manager' and row[1] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        deal_id    = row[0]
        client_id  = row[2]
        task_type  = row[3]
        task_title = row[4]

        type_labels = {'call': 'Звонок', 'meeting': 'Встреча', 'letter': 'Письмо'}
        type_label = type_labels.get(task_type, task_type)

        cur.execute("UPDATE DealTasks SET IsCompleted=1, CompletedAt=GETDATE() WHERE TaskID=?", task_id)

        now = datetime.now()

        # Событие в ленту сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'task_completed', ?, 1, ?)
        """, deal_id, f'Выполнена задача: {task_title} ({type_label})', now)

        # Дублируем событие в ленту клиента
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID, SourceDealTaskID)
            VALUES (?, ?, 'task_completed', ?, ?, ?, 1, ?, ?)
        """, client_id, session['user_id'], now,
             f'Выполнена задача по сделке',
             f'Задача "{task_title}" ({type_label}) по сделке выполнена',
             deal_id, task_id)

        conn.commit()
        conn.close()
        return jsonify({'message': 'Задача выполнена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deal-tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_deal_task(task_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT dt.DealID, d.ManagerID FROM DealTasks dt INNER JOIN Deals d ON dt.DealID=d.DealID WHERE dt.TaskID=?", task_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Задача не найдена'}), 404
        if session['role'] == 'manager' and row[1] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM DealTasks WHERE TaskID=?", task_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Задача удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Документы сделки (DealDocuments) ────────────────────────────────────────

@app.route('/api/deals/<int:deal_id>/documents', methods=['POST'])
@login_required
def upload_deal_document(deal_id):
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не найден'}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Недопустимый тип файла'}), 400
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > MAX_FILE_SIZE:
        return jsonify({'error': 'Файл слишком большой (макс. 20 МБ)'}), 400

    original_name = file.filename
    ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
    stored_name = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex

    deal_upload_dir = os.path.join(UPLOAD_DIR, 'deals', str(deal_id))
    os.makedirs(deal_upload_dir, exist_ok=True)
    file_path = os.path.join(deal_upload_dir, stored_name)
    file.save(file_path)

    try:
        conn = get_connection()
        cur = conn.cursor()

        # Получаем client_id сделки
        cur.execute("SELECT ClientID FROM Deals WHERE DealID=?", deal_id)
        deal_row = cur.fetchone()
        client_id = deal_row[0] if deal_row else None

        cur.execute("""
            INSERT INTO DealDocuments (DealID, UploadedBy, FileName, OriginalName, FileSize, MimeType)
            OUTPUT INSERTED.DocID, INSERTED.UploadedAt
            VALUES (?, ?, ?, ?, ?, ?)
        """, deal_id, session['user_id'], stored_name, original_name, size, file.content_type)
        row = cur.fetchone()
        doc_id = row[0]

        now = datetime.now()

        # Событие в ленту сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'doc_added', ?, 1, ?)
        """, deal_id, f'Добавлен документ: {original_name}', now)

        # Событие в ленту клиента (дублируем)
        if client_id:
            cur.execute("""
                INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
                VALUES (?, ?, 'doc_added', ?, ?, ?, 1, ?)
            """, client_id, session['user_id'], now,
                 f'Добавлен документ в сделку',
                 f'Документ "{original_name}" добавлен в сделку',
                 deal_id)

        conn.commit()
        conn.close()
        return jsonify({'doc_id': doc_id, 'uploaded_at': row[1].isoformat(), 'original_name': original_name}), 201
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': str(e)}), 500


@app.route('/api/deal-documents/<int:doc_id>', methods=['DELETE'])
@login_required
def delete_deal_document(doc_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT dd.DealID, dd.FileName, d.ManagerID, dd.OriginalName, d.ClientID
            FROM DealDocuments dd
            INNER JOIN Deals d ON dd.DealID=d.DealID
            WHERE dd.DocID=?
        """, doc_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Документ не найден'}), 404
        if session['role'] != 'admin' and row[2] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        deal_id = row[0]
        file_name = row[1]
        original_name = row[3]
        client_id = row[4]

        cur.execute("DELETE FROM DealDocuments WHERE DocID=?", doc_id)

        now = datetime.now()

        # Событие в ленту сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'doc_deleted', ?, 1, ?)
        """, deal_id, f'Удалён документ: {original_name}', now)

        # Событие в ленту клиента
        if client_id:
            cur.execute("""
                INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic, SourceDealID)
                VALUES (?, ?, 'doc_deleted', ?, ?, ?, 1, ?)
            """, client_id, session['user_id'], now,
                 f'Удалён документ из сделки',
                 f'Документ "{original_name}" удалён из сделки',
                 deal_id)

        conn.commit()
        conn.close()
        file_path = os.path.join(UPLOAD_DIR, 'deals', str(deal_id), file_name)
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'message': 'Документ удалён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deal-documents/<int:doc_id>/download', methods=['GET'])
@login_required
def download_deal_document(doc_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT dd.FileName, dd.OriginalName, dd.DealID FROM DealDocuments dd WHERE dd.DocID=?", doc_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Документ не найден'}), 404
        file_name = row[0]
        original_name = row[1]
        deal_id = row[2]
        conn.close()
        file_dir = os.path.join(UPLOAD_DIR, 'deals', str(deal_id))
        return send_from_directory(file_dir, file_name, as_attachment=True, download_name=original_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Автоматическое создание события при переходе между этапами
def create_deal_event(conn, cur, deal_id, event_type, description):
    """Добавляет автоматическое событие в ленту сделки."""
    try:
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description)
            VALUES (?, ?, ?)
        """, deal_id, event_type, description)
    except Exception:
        pass  # Не критично если события нет


# ═══════════════════════════════════════════════════════════════════════════════
# DOCUMENTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/clients/<int:client_id>/documents', methods=['POST'])
@login_required
def upload_document(client_id):
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не найден'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Недопустимый тип файла'}), 400

    # Проверяем размер
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > MAX_FILE_SIZE:
        return jsonify({'error': 'Файл слишком большой (максимум 20 МБ)'}), 400

    original_name = file.filename
    ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
    stored_name = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex

    client_upload_dir = os.path.join(UPLOAD_DIR, str(client_id))
    os.makedirs(client_upload_dir, exist_ok=True)

    file_path = os.path.join(client_upload_dir, stored_name)
    file.save(file_path)

    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO ClientDocuments (ClientID, UploadedBy, FileName, OriginalName, FileSize, MimeType)
            OUTPUT INSERTED.DocumentID, INSERTED.UploadedAt
            VALUES (?, ?, ?, ?, ?, ?)
        """, client_id, session['user_id'], stored_name, original_name, size, file.content_type)
        row = cur.fetchone()
        doc_id = row[0]

        now = datetime.now()

        # Событие в ленту клиента (только клиентские документы — не дублируются в сделку)
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic)
            VALUES (?, ?, 'doc_added', ?, ?, ?, 1)
        """, client_id, session['user_id'], now,
             f'Добавлен документ',
             f'Документ "{original_name}" добавлен в карточку клиента')

        conn.commit()
        conn.close()
        return jsonify({
            'document_id': doc_id,
            'uploaded_at': row[1].isoformat(),
            'file_name': stored_name,
            'original_name': original_name,
            'file_size': size,
        }), 201
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': str(e)}), 500


@app.route('/api/documents/<int:doc_id>', methods=['DELETE'])
@login_required
def delete_document(doc_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT cd.UploadedBy, cd.FileName, cd.ClientID, cd.OriginalName
            FROM ClientDocuments cd WHERE cd.DocumentID=?
        """, doc_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Документ не найден'}), 404
        if session['role'] != 'admin' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        file_name = row[1]
        client_id = row[2]
        original_name = row[3]

        cur.execute("DELETE FROM ClientDocuments WHERE DocumentID=?", doc_id)

        now = datetime.now()

        # Событие в ленту клиента (только клиентские документы)
        cur.execute("""
            INSERT INTO ClientEvents (ClientID, AuthorID, EventType, EventDate, Title, Description, IsAutomatic)
            VALUES (?, ?, 'doc_deleted', ?, ?, ?, 1)
        """, client_id, session['user_id'], now,
             f'Удалён документ',
             f'Документ "{original_name}" удалён из карточки клиента')

        conn.commit()
        conn.close()

        # Удаляем файл
        file_path = os.path.join(UPLOAD_DIR, str(client_id), file_name)
        if os.path.exists(file_path):
            os.remove(file_path)

        return jsonify({'message': 'Документ удалён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>/all-documents', methods=['GET'])
@login_required
def get_client_all_documents(client_id):
    """Документы клиента + документы из его сделок."""
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Собственные документы клиента
        cur.execute("""
            SELECT cd.DocumentID AS DocID, cd.FileName, cd.OriginalName,
                   cd.FileSize, cd.MimeType, cd.UploadedAt,
                   u.FullName AS UploaderName,
                   'client' AS DocSource, NULL AS DealID, NULL AS DealTitle
            FROM ClientDocuments cd
            INNER JOIN Users u ON cd.UploadedBy = u.UserID
            WHERE cd.ClientID = ?
            ORDER BY cd.UploadedAt DESC
        """, client_id)
        client_docs = rows_to_list(cur, cur.fetchall())

        # Документы из сделок клиента
        cur.execute("""
            SELECT dd.DocID, dd.FileName, dd.OriginalName,
                   dd.FileSize, dd.MimeType, dd.UploadedAt,
                   u.FullName AS UploaderName,
                   'deal' AS DocSource, d.DealID, d.Title AS DealTitle
            FROM DealDocuments dd
            INNER JOIN Deals d ON dd.DealID = d.DealID
            INNER JOIN Users u ON dd.UploadedBy = u.UserID
            WHERE d.ClientID = ?
            ORDER BY dd.UploadedAt DESC
        """, client_id)
        deal_docs = rows_to_list(cur, cur.fetchall())

        conn.close()
        return jsonify({'client_docs': client_docs, 'deal_docs': deal_docs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/documents/<int:doc_id>/download', methods=['GET'])
@login_required
def download_document(doc_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT cd.FileName, cd.OriginalName, cd.ClientID
            FROM ClientDocuments cd WHERE cd.DocumentID=?
        """, doc_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Документ не найден'}), 404
        file_name    = row[0]
        original_name = row[1]
        client_id    = row[2]
        conn.close()

        file_dir = os.path.join(UPLOAD_DIR, str(client_id))
        return send_from_directory(file_dir, file_name, as_attachment=True, download_name=original_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/products', methods=['GET'])
@login_required
def get_products():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT p.ProductID, p.ProductName, p.SKU, p.Unit, p.Price,
                   pc.CategoryName, pc.CategoryID, p.IsActive,
                   p.TileWidth, p.TileLength, p.TileThickness
            FROM Products p
            INNER JOIN ProductCategories pc ON p.CategoryID = pc.CategoryID
            WHERE p.IsActive = 1
            ORDER BY pc.CategoryName, p.ProductName
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# USERS / MANAGERS (Admin only)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/managers', methods=['GET'])
@login_required
def get_managers():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.UserID, u.Login, u.FullName, u.Email, u.Phone,
                   u.IsActive, u.IsPendingRole, u.CreatedAt, u.LastLoginAt,
                   (SELECT COUNT(*) FROM Clients c WHERE c.ManagerID = u.UserID) AS ClientCount,
                   ISNULL((SELECT SUM(ISNULL(d.Budget,0)) FROM Deals d
                             INNER JOIN DealStages ds ON d.StageID = ds.StageID
                             WHERE d.ManagerID = u.UserID
                             AND ds.IsCompleted = 1
                             AND d.CompletedAt IS NOT NULL
                             AND YEAR(d.CompletedAt) = YEAR(GETDATE())
                             AND MONTH(d.CompletedAt) = MONTH(GETDATE())), 0) AS MonthlySales
            FROM Users u
            INNER JOIN Roles r ON u.RoleID = r.RoleID
            WHERE r.RoleName = 'manager'
            ORDER BY u.FullName
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/managers/pending', methods=['GET'])
@admin_required
def get_pending_managers():
    """Пользователи, зарегистрированные самостоятельно — ожидают роли."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.UserID, u.Login, u.FullName, u.Email, u.Phone,
                   u.IsActive, u.IsPendingRole, u.CreatedAt
            FROM Users u
            WHERE u.IsPendingRole = 1
            ORDER BY u.CreatedAt DESC
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/managers/pending/<int:user_id>/approve', methods=['POST'])
@admin_required
def approve_pending_manager(user_id):
    """Назначить роль 'manager' зарегистрировавшемуся пользователю."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE Users
            SET RoleID = (SELECT RoleID FROM Roles WHERE RoleName = 'manager'),
                IsActive = 1,
                IsPendingRole = 0
            WHERE UserID = ? AND IsPendingRole = 1
        """, user_id)
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Пользователь не найден или уже одобрен'}), 404
        conn.commit()
        conn.close()
        return jsonify({'message': 'Роль менеджера назначена, аккаунт активирован'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/managers/pending/<int:user_id>/reject', methods=['POST'])
@admin_required
def reject_pending_manager(user_id):
    """Отклонить регистрацию (удалить пользователя)."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM Users WHERE UserID = ? AND IsPendingRole = 1", user_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Заявка отклонена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/managers', methods=['POST'])
@admin_required
def create_manager():
    data = request.get_json()
    required = ['login', 'password', 'full_name']
    for f in required:
        if not data.get(f):
            return jsonify({'error': f'Поле {f} обязательно'}), 400

    pwd_hash = hash_password(data['password'])
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO Users (RoleID, Login, PasswordHash, FullName, Email, Phone, IsActive, IsPendingRole)
            OUTPUT INSERTED.UserID
            SELECT r.RoleID, ?, ?, ?, ?, ?, 1, 0
            FROM Roles r WHERE r.RoleName = 'manager'
        """, data['login'], pwd_hash, data['full_name'],
             data.get('email'), data.get('phone'))
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'user_id': row[0], 'message': 'Менеджер создан'}), 201
    except pyodbc.IntegrityError:
        return jsonify({'error': 'Логин уже занят'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/managers/<int:user_id>/toggle', methods=['POST'])
@admin_required
def toggle_manager(user_id):
    data = request.get_json()
    is_active = data.get('is_active')
    if is_active is None:
        return jsonify({'error': 'Укажите is_active'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("EXEC sp_ToggleUserActive @UserID=?, @IsActive=?",
                    user_id, 1 if is_active else 0)
        conn.commit()
        conn.close()
        status = 'разблокирован' if is_active else 'заблокирован'
        return jsonify({'message': f'Пользователь {status}'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/managers/<int:user_id>', methods=['PUT'])
@admin_required
def update_manager(user_id):
    data = request.get_json()
    try:
        conn = get_connection()
        cur = conn.cursor()
        if data.get('password'):
            pwd_hash = hash_password(data['password'])
            cur.execute("""
                UPDATE Users SET
                    FullName     = ISNULL(?, FullName),
                    Email        = ?,
                    Phone        = ?,
                    PasswordHash = ?
                WHERE UserID = ?
            """, data.get('full_name'), data.get('email'),
                 data.get('phone'), pwd_hash, user_id)
        else:
            cur.execute("""
                UPDATE Users SET
                    FullName = ISNULL(?, FullName),
                    Email    = ?,
                    Phone    = ?
                WHERE UserID = ?
            """, data.get('full_name'), data.get('email'),
                 data.get('phone'), user_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Менеджер обновлён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/users', methods=['GET'])
@admin_required
def get_all_users():
    """Все пользователи с ролями (для управления ролями)."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.UserID, u.Login, u.FullName, u.Email, u.Phone,
                   u.IsActive, u.IsPendingRole, u.CreatedAt, u.LastLoginAt,
                   r.RoleID, r.RoleName
            FROM Users u
            LEFT JOIN Roles r ON u.RoleID = r.RoleID
            ORDER BY u.CreatedAt DESC
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/users/<int:user_id>/role', methods=['POST'])
@admin_required
def change_user_role(user_id):
    """Изменить роль пользователя (назначить manager/admin или убрать роль)."""
    data = request.get_json()
    role_name = data.get('role_name')  # 'manager', 'admin', или None/'none' для сброса
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Нельзя изменить роль себе
        if user_id == session['user_id']:
            conn.close()
            return jsonify({'error': 'Нельзя изменить собственную роль'}), 403

        if not role_name or role_name.lower() == 'none':
            # Убираем роль, деактивируем аккаунт
            cur.execute("""
                UPDATE Users SET RoleID = NULL, IsActive = 0, IsPendingRole = 1
                WHERE UserID = ?
            """, user_id)
        else:
            # Назначаем роль
            cur.execute("""
                UPDATE Users SET
                    RoleID = (SELECT RoleID FROM Roles WHERE RoleName = ?),
                    IsActive = 1,
                    IsPendingRole = 0
                WHERE UserID = ?
            """, role_name.lower(), user_id)

        conn.commit()
        conn.close()
        return jsonify({'message': f'Роль обновлена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# ПЛАНЫ ПРОДАЖ И КВОТЫ
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/sales-plans', methods=['GET'])
@login_required
def get_sales_plans():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT sp.PlanID, sp.PlanYear, sp.PlanMonth, sp.TargetAmount,
                   sp.CreatedAt, u.FullName AS CreatedByName
            FROM SalesPlans sp
            INNER JOIN Users u ON sp.CreatedBy = u.UserID
            ORDER BY sp.PlanYear DESC, sp.PlanMonth DESC
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sales-plans', methods=['POST'])
@admin_required
def create_sales_plan():
    data = request.get_json()
    if not data.get('plan_year') or not data.get('plan_month') or not data.get('target_amount'):
        return jsonify({'error': 'Год, месяц и целевая сумма обязательны'}), 400
    try:
        ta = float(data['target_amount'])
        if ta < 1:
            return jsonify({'error': 'Целевая сумма не может быть меньше 1'}), 400
        if ta > 999_999_999:
            return jsonify({'error': 'Целевая сумма не может превышать 999 999 999'}), 400
    except (TypeError, ValueError):
        return jsonify({'error': 'Некорректная целевая сумма'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            MERGE SalesPlans AS target
            USING (VALUES (?, ?, ?, ?)) AS source(PlanYear, PlanMonth, TargetAmount, CreatedBy)
            ON target.PlanYear = source.PlanYear AND target.PlanMonth = source.PlanMonth
            WHEN MATCHED THEN
                UPDATE SET TargetAmount = source.TargetAmount
            WHEN NOT MATCHED THEN
                INSERT (PlanYear, PlanMonth, TargetAmount, CreatedBy)
                VALUES (source.PlanYear, source.PlanMonth, source.TargetAmount, source.CreatedBy);
        """, data['plan_year'], data['plan_month'], data['target_amount'], session['user_id'])
        conn.commit()
        conn.close()
        return jsonify({'message': 'План продаж сохранён'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sales-plans/<int:plan_id>', methods=['DELETE'])
@admin_required
def delete_sales_plan(plan_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM SalesPlans WHERE PlanID=?", plan_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'План удалён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sales-plans/current', methods=['GET'])
@login_required
def get_current_sales_plan():
    """Получить текущий план продаж с фактическим прогрессом для текущего пользователя."""
    now = datetime.now()
    manager_id = session['user_id']
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT sp.PlanID, sp.PlanYear, sp.PlanMonth, sp.TargetAmount
            FROM SalesPlans sp
            WHERE sp.PlanYear = ? AND sp.PlanMonth = ?
        """, now.year, now.month)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify(None)
        plan = row_to_dict(cur, row)

        # Фактические продажи за месяц: завершённые сделки ТОЛЬКО текущего пользователя
        cur.execute("""
            SELECT ISNULL(SUM(ISNULL(d.Budget,0)), 0) AS ActualAmount
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ManagerID = ?
              AND ds.IsCompleted = 1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt) = ? AND MONTH(d.CompletedAt) = ?
        """, manager_id, now.year, now.month)
        actual = cur.fetchone()
        plan['ActualAmount'] = float(actual[0]) if actual else 0.0
        plan['ProgressPct']  = round(plan['ActualAmount'] / float(plan['TargetAmount']) * 100, 1) if plan['TargetAmount'] else 0.0

        conn.close()
        return jsonify(plan)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sales-plans/company', methods=['GET'])
@login_required
def get_company_sales_plan():
    """Получить план продаж с фактическим прогрессом по ВСЕМ пользователям (завершённые сделки)."""
    now = datetime.now()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT sp.PlanID, sp.PlanYear, sp.PlanMonth, sp.TargetAmount
            FROM SalesPlans sp
            WHERE sp.PlanYear = ? AND sp.PlanMonth = ?
        """, now.year, now.month)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify(None)
        plan = row_to_dict(cur, row)

        # Фактические продажи за месяц: завершённые сделки ВСЕХ пользователей
        cur.execute("""
            SELECT ISNULL(SUM(ISNULL(d.Budget,0)), 0) AS ActualAmount
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE ds.IsCompleted = 1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt) = ? AND MONTH(d.CompletedAt) = ?
        """, now.year, now.month)
        actual = cur.fetchone()
        plan['ActualAmount'] = float(actual[0]) if actual else 0.0
        plan['ProgressPct']  = round(plan['ActualAmount'] / float(plan['TargetAmount']) * 100, 1) if plan['TargetAmount'] else 0.0

        conn.close()
        return jsonify(plan)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/quotas', methods=['GET'])
@login_required
def get_quotas():
    manager_id = request.args.get('manager_id') or (session['user_id'] if session['role'] == 'manager' else None)
    try:
        conn = get_connection()
        cur = conn.cursor()
        query = """
            SELECT mq.QuotaID, mq.ManagerID, mq.QuotaYear, mq.QuotaMonth,
                   mq.QuotaType, mq.TargetValue, mq.CreatedAt,
                   u.FullName AS ManagerName
            FROM ManagerQuotas mq
            INNER JOIN Users u ON mq.ManagerID = u.UserID
            WHERE 1=1
        """
        params = []
        if manager_id:
            query += " AND mq.ManagerID = ?"
            params.append(manager_id)
        query += " ORDER BY mq.QuotaYear DESC, mq.QuotaMonth DESC, mq.ManagerID"
        cur.execute(query, *params)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/quotas', methods=['POST'])
@admin_required
def create_quota():
    data = request.get_json()
    required = ['manager_id', 'quota_year', 'quota_month', 'quota_type', 'target_value']
    for f in required:
        if data.get(f) is None:
            return jsonify({'error': f'Поле {f} обязательно'}), 400
    # Ограничение значения
    quota_type = data.get('quota_type', '')
    try:
        tv = float(data['target_value'])
        if tv < 1:
            return jsonify({'error': 'Целевое значение не может быть меньше 1'}), 400
        if quota_type == 'plan_percent':
            if tv > 100:
                return jsonify({'error': 'Для типа "Выполнение плана (%)" значение не может превышать 100'}), 400
        else:
            if tv > 999_999_999:
                return jsonify({'error': 'Целевое значение слишком большое (максимум 999 999 999)'}), 400
    except (TypeError, ValueError):
        return jsonify({'error': 'Некорректное значение квоты'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        # Если тип plan_percent — проверить наличие плана продаж на этот месяц
        if quota_type == 'plan_percent':
            cur.execute(
                "SELECT COUNT(*) FROM SalesPlans WHERE PlanYear=? AND PlanMonth=?",
                data['quota_year'], data['quota_month']
            )
            cnt = cur.fetchone()[0]
            if not cnt:
                conn.close()
                month_names = ['январь','февраль','март','апрель','май','июнь',
                               'июль','август','сентябрь','октябрь','ноябрь','декабрь']
                m_name = month_names[int(data['quota_month']) - 1] if 1 <= int(data['quota_month']) <= 12 else data['quota_month']
                return jsonify({'error': f'Нельзя добавить квоту "Выполнение плана (%)" — на {m_name} {data["quota_year"]} не создан план продаж'}), 400
        cur.execute("""
            MERGE ManagerQuotas AS target
            USING (VALUES (?, ?, ?, ?, ?, ?)) AS source(ManagerID, QuotaYear, QuotaMonth, QuotaType, TargetValue, CreatedBy)
            ON target.ManagerID = source.ManagerID
               AND target.QuotaYear = source.QuotaYear
               AND target.QuotaMonth = source.QuotaMonth
               AND target.QuotaType = source.QuotaType
            WHEN MATCHED THEN
                UPDATE SET TargetValue = source.TargetValue
            WHEN NOT MATCHED THEN
                INSERT (ManagerID, QuotaYear, QuotaMonth, QuotaType, TargetValue, CreatedBy)
                VALUES (source.ManagerID, source.QuotaYear, source.QuotaMonth,
                        source.QuotaType, source.TargetValue, source.CreatedBy);
        """, data['manager_id'], data['quota_year'], data['quota_month'],
             data['quota_type'], data['target_value'], session['user_id'])
        conn.commit()
        conn.close()
        return jsonify({'message': 'Квота сохранена'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/quotas/<int:quota_id>', methods=['DELETE'])
@admin_required
def delete_quota(quota_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM ManagerQuotas WHERE QuotaID=?", quota_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Квота удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/quotas/progress', methods=['GET'])
@login_required
def get_quota_progress():
    """Получить квоты менеджера с фактическим прогрессом."""
    manager_id = int(request.args.get('manager_id') or session['user_id'])
    year  = int(request.args.get('year',  datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT mq.QuotaID, mq.QuotaType, mq.TargetValue,
                   mq.QuotaYear, mq.QuotaMonth
            FROM ManagerQuotas mq
            WHERE mq.ManagerID = ? AND mq.QuotaYear = ? AND mq.QuotaMonth = ?
        """, manager_id, year, month)
        quotas = rows_to_list(cur, cur.fetchall())

        # Продажи за месяц: завершённые сделки (IsCompleted=1), CompletedAt в этом месяце
        cur.execute("""
            SELECT ISNULL(SUM(ISNULL(d.Budget,0)),0)
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ManagerID = ?
              AND ds.IsCompleted = 1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt) = ?
              AND MONTH(d.CompletedAt) = ?
        """, manager_id, year, month)
        sales_row = cur.fetchone()
        fact_sales = float(sales_row[0] or 0) if sales_row else 0.0

        # Количество завершённых сделок за этот месяц
        cur.execute("""
            SELECT COUNT(*)
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.ManagerID = ?
              AND ds.IsCompleted = 1
              AND d.CompletedAt IS NOT NULL
              AND YEAR(d.CompletedAt) = ?
              AND MONTH(d.CompletedAt) = ?
        """, manager_id, year, month)
        deals_row = cur.fetchone()
        fact_deals = int(deals_row[0] or 0) if deals_row else 0

        cur.execute("""
            SELECT COUNT(*) AS NewClients
            FROM Clients
            WHERE ManagerID = ?
              AND YEAR(CreatedAt) = ?
              AND MONTH(CreatedAt) = ?
        """, manager_id, year, month)
        nc_row = cur.fetchone()
        fact_new_clients = int(nc_row[0] or 0) if nc_row else 0

        # Получаем план за месяц для расчёта %
        cur.execute("""
            SELECT TargetAmount FROM SalesPlans
            WHERE PlanYear = ? AND PlanMonth = ?
        """, year, month)
        plan_row = cur.fetchone()
        plan_target = float(plan_row[0]) if (plan_row and plan_row[0] is not None) else None
        plan_pct = round(fact_sales / plan_target * 100, 1) if plan_target else 0.0

        result = []
        for q in quotas:
            qt    = q['QuotaType']
            target = float(q['TargetValue'])
            actual = 0.0
            if qt == 'sales_amount':
                actual = fact_sales
            elif qt == 'deals_count':
                actual = float(fact_deals)
            elif qt == 'new_clients':
                actual = float(fact_new_clients)
            elif qt == 'plan_percent':
                actual = plan_pct
            pct = round(actual / target * 100, 1) if target > 0 else 0.0
            result.append({
                **q,
                'ActualValue': actual,
                'ProgressPct': min(pct, 100.0),
            })

        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# СПРАВОЧНИКИ
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/references', methods=['GET'])
@login_required
def get_references():
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT ClientTypeID, TypeName FROM ClientTypes ORDER BY ClientTypeID")
        client_types = rows_to_list(cur, cur.fetchall())

        cur.execute("SELECT LeadSourceID, SourceName FROM LeadSources ORDER BY LeadSourceID")
        lead_sources = rows_to_list(cur, cur.fetchall())

        cur.execute("SELECT StatusID, StatusName FROM ClientStatuses ORDER BY StatusID")
        statuses = rows_to_list(cur, cur.fetchall())

        cur.execute("SELECT CategoryID, CategoryName FROM ProductCategories ORDER BY CategoryID")
        categories = rows_to_list(cur, cur.fetchall())

        cur.execute("SELECT StageID, StageName, StageOrder, IsCompleted FROM DealStages ORDER BY StageOrder")
        deal_stages = rows_to_list(cur, cur.fetchall())

        conn.close()
        return jsonify({
            'client_types': client_types,
            'lead_sources': lead_sources,
            'statuses':     statuses,
            'categories':   categories,
            'deal_stages':  deal_stages,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# HEALTH CHECK
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/health', methods=['GET'])
def health():
    try:
        conn = get_connection()
        conn.close()
        return jsonify({'status': 'ok', 'db': 'connected'})
    except Exception as e:
        return jsonify({'status': 'error', 'db': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# КОММЕРЧЕСКИЕ ПРЕДЛОЖЕНИЯ (КП)
# ═══════════════════════════════════════════════════════════════════════════════

def generate_offer_number(cur, year):
    cur.execute("SELECT SeqValue FROM OfferNumberSequence WHERE SeqYear=?", year)
    row = cur.fetchone()
    if row:
        new_val = row[0] + 1
        cur.execute("UPDATE OfferNumberSequence SET SeqValue=? WHERE SeqYear=?", new_val, year)
    else:
        new_val = 1
        cur.execute("INSERT INTO OfferNumberSequence (SeqYear, SeqValue) VALUES (?,?)", year, new_val)
    return f"КП-{year}-{new_val:05d}"


def _update_expired_offers(cur, offer_ids=None):
    """Обновляет статус просроченных КП."""
    today = date.today().isoformat()
    if offer_ids:
        placeholders = ','.join(['?' for _ in offer_ids])
        cur.execute(f"""
            UPDATE CommercialOffers SET Status='expired'
            WHERE Status='active' AND ExpiresAt < ? AND OfferID IN ({placeholders})
        """, today, *offer_ids)
    else:
        cur.execute("""
            UPDATE CommercialOffers SET Status='expired'
            WHERE Status='active' AND ExpiresAt < ?
        """, today)


@app.route('/api/offers', methods=['GET'])
@login_required
def get_offers():
    get_all = request.args.get('all') == '1'
    get_mine = request.args.get('my') == '1'
    client_id_filter = request.args.get('client_id')
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Обновляем просроченные КП
        _update_expired_offers(cur)
        conn.commit()

        query = """
            SELECT co.OfferID, co.OfferNumber, co.ClientID, co.DealID,
                   co.CreatedBy, co.CreatedAt, co.ExpiresAt, co.Status, co.Notes, co.UsedAt,
                   c.FullName AS ClientName, c.CompanyName AS ClientCompany,
                   c.ManagerID AS ClientManagerID,
                   u.FullName AS CreatedByName,
                   d.Title AS DealTitle
            FROM CommercialOffers co
            INNER JOIN Clients c ON co.ClientID = c.ClientID
            INNER JOIN Users u ON co.CreatedBy = u.UserID
            LEFT JOIN Deals d ON co.DealID = d.DealID
            WHERE 1=1
        """
        params = []

        if get_mine:
            # Показывать только КП для клиентов закреплённых за текущим пользователем
            query += " AND c.ManagerID = ?"
            params.append(session['user_id'])
        elif session['role'] == 'manager':
            # Менеджер видит только КП своих клиентов
            query += " AND c.ManagerID = ?"
            params.append(session['user_id'])
        elif not get_all:
            # admin без параметров — показывать все (поведение по умолчанию для админа)
            pass

        if client_id_filter:
            query += " AND co.ClientID = ?"
            params.append(client_id_filter)

        query += " ORDER BY co.CreatedAt DESC"
        cur.execute(query, *params)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offers', methods=['POST'])
@login_required
def create_offer():
    data = request.get_json()
    if not data.get('client_id') or not data.get('expires_at'):
        return jsonify({'error': 'client_id и expires_at обязательны'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Проверяем доступ к клиенту
        if session['role'] == 'manager':
            cur.execute("SELECT ManagerID FROM Clients WHERE ClientID=?", data['client_id'])
            row = cur.fetchone()
            if not row or row[0] != session['user_id']:
                conn.close()
                return jsonify({'error': 'Нет доступа к этому клиенту'}), 403

        year = datetime.now().year
        offer_number = generate_offer_number(cur, year)

        cur.execute("""
            INSERT INTO CommercialOffers (OfferNumber, ClientID, DealID, CreatedBy, ExpiresAt, Status, Notes)
            OUTPUT INSERTED.OfferID
            VALUES (?, ?, ?, ?, ?, 'active', ?)
        """, offer_number, data['client_id'],
             data.get('deal_id') or None,
             session['user_id'],
             data['expires_at'],
             data.get('notes') or None)
        row = cur.fetchone()
        offer_id = row[0]
        conn.commit()
        conn.close()
        return jsonify({'offer_id': offer_id, 'offer_number': offer_number, 'message': 'КП создано'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offers/<int:offer_id>', methods=['GET'])
@login_required
def get_offer(offer_id):
    try:
        conn = get_connection()
        cur = conn.cursor()

        _update_expired_offers(cur, [offer_id])
        conn.commit()

        cur.execute("""
            SELECT co.OfferID, co.OfferNumber, co.ClientID, co.DealID,
                   co.CreatedBy, co.CreatedAt, co.ExpiresAt, co.Status, co.Notes, co.UsedAt,
                   c.FullName AS ClientName, c.CompanyName AS ClientCompany,
                   u.FullName AS CreatedByName,
                   d.Title AS DealTitle
            FROM CommercialOffers co
            INNER JOIN Clients c ON co.ClientID = c.ClientID
            INNER JOIN Users u ON co.CreatedBy = u.UserID
            LEFT JOIN Deals d ON co.DealID = d.DealID
            WHERE co.OfferID = ?
        """, offer_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'КП не найдено'}), 404
        offer = row_to_dict(cur, row)

        if session['role'] == 'manager' and offer['CreatedBy'] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        # Items
        cur.execute("""
            SELECT coi.ItemID, coi.ProductID, p.ProductName, p.SKU, p.Unit,
                   p.Price AS BasePrice, coi.DiscountPct,
                   ROUND(p.Price * (1 - coi.DiscountPct/100.0), 2) AS DiscountedPrice,
                   p.TileWidth, p.TileLength, p.TileThickness
            FROM CommercialOfferItems coi
            INNER JOIN Products p ON coi.ProductID = p.ProductID
            WHERE coi.OfferID = ?
            ORDER BY coi.ItemID
        """, offer_id)
        offer['items'] = rows_to_list(cur, cur.fetchall())

        conn.close()
        return jsonify(offer)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offers/<int:offer_id>', methods=['PUT'])
@login_required
def update_offer(offer_id):
    data = request.get_json()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT CreatedBy, Status FROM CommercialOffers WHERE OfferID=?", offer_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'КП не найдено'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        cur.execute("""
            UPDATE CommercialOffers SET
                Notes = ?,
                ExpiresAt = ISNULL(?, ExpiresAt)
            WHERE OfferID = ?
        """, data.get('notes'), data.get('expires_at') or None, offer_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'КП обновлено'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offers/<int:offer_id>', methods=['DELETE'])
@login_required
def delete_offer(offer_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT CreatedBy FROM CommercialOffers WHERE OfferID=?", offer_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'КП не найдено'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM CommercialOffers WHERE OfferID=?", offer_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'КП удалено'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offers/<int:offer_id>/items', methods=['GET'])
@login_required
def get_offer_items(offer_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT coi.ItemID, coi.ProductID, p.ProductName, p.SKU, p.Unit,
                   p.Price AS BasePrice, coi.DiscountPct,
                   ROUND(p.Price * (1 - coi.DiscountPct/100.0), 2) AS DiscountedPrice,
                   p.TileWidth, p.TileLength, p.TileThickness
            FROM CommercialOfferItems coi
            INNER JOIN Products p ON coi.ProductID = p.ProductID
            WHERE coi.OfferID = ?
            ORDER BY coi.ItemID
        """, offer_id)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offers/<int:offer_id>/items', methods=['POST'])
@login_required
def add_offer_item(offer_id):
    data = request.get_json()
    if not data.get('product_id'):
        return jsonify({'error': 'product_id обязателен'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT CreatedBy FROM CommercialOffers WHERE OfferID=?", offer_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'КП не найдено'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403

        discount = float(data.get('discount_pct') or 0)
        if discount < 0 or discount > 100:
            conn.close()
            return jsonify({'error': 'Скидка должна быть от 0 до 100%'}), 400

        # Upsert
        cur.execute("SELECT ItemID FROM CommercialOfferItems WHERE OfferID=? AND ProductID=?",
                    offer_id, data['product_id'])
        existing = cur.fetchone()
        if existing:
            cur.execute("UPDATE CommercialOfferItems SET DiscountPct=? WHERE ItemID=?",
                        discount, existing[0])
        else:
            cur.execute("""
                INSERT INTO CommercialOfferItems (OfferID, ProductID, DiscountPct)
                VALUES (?, ?, ?)
            """, offer_id, data['product_id'], discount)

        conn.commit()
        conn.close()
        return jsonify({'message': 'Позиция добавлена в КП'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offer-items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_offer_item(item_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT coi.OfferID, co.CreatedBy
            FROM CommercialOfferItems coi
            INNER JOIN CommercialOffers co ON coi.OfferID = co.OfferID
            WHERE coi.ItemID = ?
        """, item_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Позиция не найдена'}), 404
        if session['role'] == 'manager' and row[1] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        cur.execute("DELETE FROM CommercialOfferItems WHERE ItemID=?", item_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Позиция удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>/apply-offer', methods=['POST'])
@login_required
def apply_offer_to_deal(deal_id):
    data = request.get_json()
    offer_id = data.get('offer_id')
    if not offer_id:
        return jsonify({'error': 'offer_id обязателен'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Проверяем сделку
        cur.execute("""
            SELECT d.ManagerID, d.ClientID, ds.StageOrder
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.DealID = ?
        """, deal_id)
        deal_row = cur.fetchone()
        if not deal_row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        if session['role'] == 'manager' and deal_row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        if deal_row[2] >= 5:
            conn.close()
            return jsonify({'error': "Нельзя применять КП начиная с этапа 'Счёт выставлен'"}), 400

        deal_client_id = deal_row[1]

        # Проверяем КП
        _update_expired_offers(cur, [offer_id])
        cur.execute("""
            SELECT OfferID, Status, ExpiresAt, ClientID
            FROM CommercialOffers WHERE OfferID=?
        """, offer_id)
        offer_row = cur.fetchone()
        if not offer_row:
            conn.close()
            return jsonify({'error': 'КП не найдено'}), 404
        if offer_row[1] != 'active':
            conn.close()
            return jsonify({'error': 'КП не активно'}), 400
        if offer_row[3] != deal_client_id:
            conn.close()
            return jsonify({'error': 'КП не принадлежит клиенту этой сделки'}), 400

        expires = offer_row[2]
        if isinstance(expires, str):
            expires = date.fromisoformat(expires)
        if expires < date.today():
            conn.close()
            return jsonify({'error': 'Срок действия КП истёк'}), 400

        # Применяем КП
        cur.execute("UPDATE Deals SET AppliedOfferID=?, UpdatedAt=GETDATE() WHERE DealID=?",
                    offer_id, deal_id)

        # Получаем позиции КП для пересчёта цен
        cur.execute("""
            SELECT coi.OfferID, coi.ProductID, p.Price AS BasePrice, coi.DiscountPct,
                   ROUND(p.Price * (1 - coi.DiscountPct/100.0), 2) AS DiscountedPrice
            FROM CommercialOfferItems coi
            INNER JOIN Products p ON coi.ProductID = p.ProductID
            WHERE coi.OfferID = ?
        """, offer_id)
        offer_items = rows_to_list(cur, cur.fetchall())

        # Обновляем цены уже добавленных позиций в составе заказа сделки
        # LineTotal — вычисляемая колонка (Quantity * UnitPrice), обновляется автоматически
        for oi in offer_items:
            cur.execute("""
                UPDATE DealItems SET UnitPrice = ?
                WHERE DealID = ? AND ProductID = ?
            """, oi['DiscountedPrice'], deal_id, oi['ProductID'])

        # Пересчитываем общий бюджет сделки
        cur.execute("""
            UPDATE Deals SET
                Budget = (SELECT ISNULL(SUM(LineTotal), 0) FROM DealItems WHERE DealID = ?),
                UpdatedAt = GETDATE()
            WHERE DealID = ?
        """, deal_id, deal_id)

        # Событие в ленту сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'offer_applied', ?, 1, GETDATE())
        """, deal_id, f'Применено КП: {offer_items[0]["OfferID"] if offer_items else offer_id}')

        conn.commit()
        conn.close()
        return jsonify({'message': 'КП применено', 'offer_items': offer_items})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/deals/<int:deal_id>/apply-offer', methods=['DELETE'])
@login_required
def remove_offer_from_deal(deal_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT d.ManagerID, ds.StageOrder
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID = ds.StageID
            WHERE d.DealID = ?
        """, deal_id)
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Сделка не найдена'}), 404
        if session['role'] == 'manager' and row[0] != session['user_id']:
            conn.close()
            return jsonify({'error': 'Нет доступа'}), 403
        if row[1] >= 5:
            conn.close()
            return jsonify({'error': "Нельзя изменять КП начиная с этапа 'Счёт выставлен'"}), 400

        # Получаем товары с ценами из КП чтобы потом восстановить базовые цены
        cur.execute("""
            SELECT di.ProductID, p.Price AS BasePrice
            FROM DealItems di
            INNER JOIN Products p ON di.ProductID = p.ProductID
            WHERE di.DealID = ?
        """, deal_id)
        items_to_restore = rows_to_list(cur, cur.fetchall())

        cur.execute("UPDATE Deals SET AppliedOfferID=NULL, UpdatedAt=GETDATE() WHERE DealID=?", deal_id)

        # Восстанавливаем базовые цены для всех позиций заказа
        # LineTotal — вычисляемая колонка (Quantity * UnitPrice), обновляется автоматически
        for item in items_to_restore:
            cur.execute("""
                UPDATE DealItems SET UnitPrice = ?
                WHERE DealID = ? AND ProductID = ?
            """, item['BasePrice'], deal_id, item['ProductID'])

        # Пересчитываем общий бюджет сделки
        cur.execute("""
            UPDATE Deals SET
                Budget = (SELECT ISNULL(SUM(LineTotal), 0) FROM DealItems WHERE DealID = ?),
                UpdatedAt = GETDATE()
            WHERE DealID = ?
        """, deal_id, deal_id)

        # Событие в ленту сделки
        cur.execute("""
            INSERT INTO DealEvents (DealID, EventType, Description, IsAutomatic, EventDate)
            VALUES (?, 'offer_removed', 'КП откреплено от сделки', 1, GETDATE())
        """, deal_id)

        conn.commit()
        conn.close()
        return jsonify({'message': 'КП откреплено'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# ОТЧЁТЫ
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/reports/managers', methods=['GET'])
@admin_required
def report_managers():
    from_date = request.args.get('from')
    to_date   = request.args.get('to')
    try:
        conn = get_connection()
        cur = conn.cursor()

        if from_date and to_date:
            to_date_full = to_date + ' 23:59:59'
            cur.execute("""
                SELECT u.UserID, u.FullName AS ManagerName,
                       (SELECT COUNT(DISTINCT d2.DealID) FROM Deals d2
                        INNER JOIN DealStages ds2 ON d2.StageID = ds2.StageID
                        WHERE d2.ManagerID = u.UserID AND ds2.IsCompleted = 1
                        AND d2.CompletedAt >= ? AND d2.CompletedAt <= ?) AS DealsCount,
                       (SELECT ISNULL(SUM(d2.Budget), 0) FROM Deals d2
                        INNER JOIN DealStages ds2 ON d2.StageID = ds2.StageID
                        WHERE d2.ManagerID = u.UserID AND ds2.IsCompleted = 1
                        AND d2.CompletedAt >= ? AND d2.CompletedAt <= ?) AS TotalSales,
                       (SELECT COUNT(*) FROM Clients c2 WHERE c2.ManagerID=u.UserID
                        AND c2.CreatedAt >= ? AND c2.CreatedAt <= ?) AS NewClientsCount
                FROM Users u
                INNER JOIN Roles r ON u.RoleID = r.RoleID
                WHERE r.RoleName = 'manager' AND u.IsActive = 1
            """, from_date, to_date_full, from_date, to_date_full, from_date, to_date_full)
        else:
            cur.execute("""
                SELECT u.UserID, u.FullName AS ManagerName,
                       (SELECT COUNT(DISTINCT d2.DealID) FROM Deals d2
                        INNER JOIN DealStages ds2 ON d2.StageID = ds2.StageID
                        WHERE d2.ManagerID = u.UserID AND ds2.IsCompleted = 1) AS DealsCount,
                       (SELECT ISNULL(SUM(d2.Budget), 0) FROM Deals d2
                        INNER JOIN DealStages ds2 ON d2.StageID = ds2.StageID
                        WHERE d2.ManagerID = u.UserID AND ds2.IsCompleted = 1) AS TotalSales,
                       (SELECT COUNT(*) FROM Clients c2 WHERE c2.ManagerID=u.UserID) AS NewClientsCount
                FROM Users u
                INNER JOIN Roles r ON u.RoleID = r.RoleID
                WHERE r.RoleName = 'manager' AND u.IsActive = 1
            """)

        rows = cur.fetchall()
        conn.close()
        # Рассчитываем AvgDealValue после получения данных
        result = rows_to_list(cur, rows)
        for r in result:
            deals_count = r.get('DealsCount', 0) or 0
            total_sales = r.get('TotalSales', 0) or 0
            r['AvgDealValue'] = round(total_sales / deals_count, 2) if deals_count > 0 else 0
        # Сортируем по сумме продаж (по убыванию) для правильного рейтинга
        result.sort(key=lambda x: x.get('TotalSales', 0) or 0, reverse=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/reports/sources', methods=['GET'])
@admin_required
def report_sources():
    from_date = request.args.get('from')
    to_date   = request.args.get('to')
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        date_filter = ""
        if from_date and to_date:
            date_filter = " AND c.CreatedAt >= ? AND c.CreatedAt <= ?"
            params = [from_date, to_date + ' 23:59:59']
        cur.execute(f"""
            SELECT ls.SourceName, COUNT(c.ClientID) AS ClientsCount
            FROM LeadSources ls
            LEFT JOIN Clients c ON c.LeadSourceID = ls.LeadSourceID{date_filter}
            GROUP BY ls.LeadSourceID, ls.SourceName
            ORDER BY ClientsCount DESC
        """, *params)
        rows = cur.fetchall()
        conn.close()
        data = rows_to_list(cur, rows)
        total = sum(r['ClientsCount'] for r in data)
        for r in data:
            r['Percentage'] = round(r['ClientsCount'] / total * 100, 1) if total else 0
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/reports/sales', methods=['GET'])
@admin_required
def report_sales():
    from_date = request.args.get('from', date(date.today().year, 1, 1).isoformat())
    to_date   = request.args.get('to', date.today().isoformat())
    period    = request.args.get('period', 'month')
    try:
        conn = get_connection()
        cur = conn.cursor()
        if period == 'day':
            group_by = "CAST(d.CompletedAt AS DATE)"
            label_expr = "CONVERT(NVARCHAR,CAST(d.CompletedAt AS DATE),103)"
        elif period == 'week':
            group_by = "DATEPART(YEAR,d.CompletedAt), DATEPART(WEEK,d.CompletedAt)"
            label_expr = "CONCAT(DATEPART(YEAR,d.CompletedAt),N'-W',DATEPART(WEEK,d.CompletedAt))"
        else:
            group_by = "YEAR(d.CompletedAt), MONTH(d.CompletedAt)"
            label_expr = "CONCAT(YEAR(d.CompletedAt),N'-',RIGHT(N'0'+CAST(MONTH(d.CompletedAt) AS NVARCHAR),2))"

        cur.execute(f"""
            SELECT {label_expr} AS PeriodLabel, ISNULL(SUM(d.Budget),0) AS SalesAmount, COUNT(d.DealID) AS DealsCount
            FROM Deals d
            INNER JOIN DealStages ds ON d.StageID=ds.StageID
            WHERE ds.IsCompleted=1 AND d.CompletedAt IS NOT NULL
              AND d.CompletedAt >= ? AND d.CompletedAt <= ?
            GROUP BY {group_by}
            ORDER BY {group_by}
        """, from_date, to_date + ' 23:59:59')
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/reports/export', methods=['GET'])
@admin_required
def export_report():
    report_type = request.args.get('type', 'managers')
    from_date   = request.args.get('from')
    to_date     = request.args.get('to')
    fmt_param   = request.args.get('format', 'xlsx')

    from flask import Response

    # Collect data via internal call
    if report_type == 'managers':
        resp = report_managers()
        if hasattr(resp, 'get_json'):
            data = resp.get_json()
        else:
            data = []
        headers_list = ['Менеджер', 'Сделок', 'Сумма продаж (BYN)', 'Средний чек (BYN)', 'Новых клиентов']
        rows_data = [[r.get('ManagerName',''), r.get('DealsCount',0),
                      round(float(r.get('TotalSales',0)),2),
                      round(float(r.get('AvgDealValue',0)),2),
                      r.get('NewClientsCount',0)] for r in data]
        title = 'Эффективность менеджеров'
    elif report_type == 'sources':
        resp = report_sources()
        if hasattr(resp, 'get_json'):
            data = resp.get_json()
        else:
            data = []
        headers_list = ['Источник', 'Клиентов', '% от общего']
        rows_data = [[r.get('SourceName',''), r.get('ClientsCount',0), r.get('Percentage',0)] for r in data]
        title = 'Источники клиентов'
    else:
        resp = report_sales()
        if hasattr(resp, 'get_json'):
            data = resp.get_json()
        else:
            data = []
        headers_list = ['Период', 'Сумма продаж (BYN)', 'Сделок']
        rows_data = [[r.get('PeriodLabel',''), round(float(r.get('SalesAmount',0)),2), r.get('DealsCount',0)] for r in data]
        title = 'Динамика продаж'

    if fmt_param == 'xlsx' and OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        ws.append(headers_list)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows_data:
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"report_{report_type}.xlsx"
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    else:
        # HTML download (for PDF/DOCX we generate nice HTML)
        rows_html = ''.join(
            '<tr>' + ''.join(f'<td style="border:1px solid #e2e8f0;padding:8px">{v}</td>' for v in row) + '</tr>'
            for row in rows_data
        )
        headers_html = ''.join(
            f'<th style="border:1px solid #e2e8f0;padding:8px;background:#1e40af;color:#fff;text-align:left">{h}</th>'
            for h in headers_list
        )
        period_str = f'{from_date} — {to_date}' if from_date and to_date else 'Все время'
        html = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">
<title>{title}</title>
<style>@media print{{@page{{margin:2cm}}}}body{{font-family:Arial,sans-serif;margin:20px}}
table{{border-collapse:collapse;width:100%}}h1{{color:#1e40af}}</style>
</head><body>
<h1>{title}</h1><p style="color:#64748b">Период: {period_str}</p>
<table><thead><tr>{headers_html}</tr></thead><tbody>{rows_html}</tbody></table>
<p style="color:#64748b;font-size:11px;margin-top:16px">Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
</body></html>"""
        filename = f"report_{report_type}.html"
        return Response(html, mimetype='text/html',
                       headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# СПРАВОЧНИКИ (CRUD для администратора)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/ref/lead-sources', methods=['GET'])
@login_required
def ref_get_lead_sources():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT LeadSourceID, SourceName FROM LeadSources ORDER BY LeadSourceID")
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/lead-sources', methods=['POST'])
@admin_required
def ref_create_lead_source():
    data = request.get_json()
    name = data.get('name', '')
    if not name.strip():
        return jsonify({'error': 'Название обязательно'}), 400
    if len(name) > 100:
        return jsonify({'error': 'Название слишком длинное (максимум 100 символов)'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO LeadSources (SourceName) OUTPUT INSERTED.LeadSourceID VALUES (?)", name)
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'id': row[0], 'message': 'Источник создан'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/lead-sources/<int:src_id>', methods=['PUT'])
@admin_required
def ref_update_lead_source(src_id):
    data = request.get_json()
    name = data.get('name', '')
    if not name.strip():
        return jsonify({'error': 'Название обязательно'}), 400
    if len(name) > 100:
        return jsonify({'error': 'Название слишком длинное (максимум 100 символов)'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("UPDATE LeadSources SET SourceName=? WHERE LeadSourceID=?", name, src_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Источник обновлён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/lead-sources/<int:src_id>', methods=['DELETE'])
@admin_required
def ref_delete_lead_source(src_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM Clients WHERE LeadSourceID=?", src_id)
        cnt = cur.fetchone()[0]
        if cnt > 0:
            conn.close()
            return jsonify({'error': f'Нельзя удалить: используется в {cnt} клиентах'}), 409
        cur.execute("DELETE FROM LeadSources WHERE LeadSourceID=?", src_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Источник удалён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/client-statuses', methods=['GET'])
@login_required
def ref_get_client_statuses():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT StatusID, StatusName FROM ClientStatuses ORDER BY StatusID")
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/client-statuses', methods=['POST'])
@admin_required
def ref_create_client_status():
    data = request.get_json()
    if not data.get('name'):
        return jsonify({'error': 'Название обязательно'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO ClientStatuses (StatusName) OUTPUT INSERTED.StatusID VALUES (?)", data['name'])
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'id': row[0], 'message': 'Статус создан'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/client-statuses/<int:status_id>', methods=['PUT'])
@admin_required
def ref_update_client_status(status_id):
    data = request.get_json()
    if not data.get('name'):
        return jsonify({'error': 'Название обязательно'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("UPDATE ClientStatuses SET StatusName=? WHERE StatusID=?", data['name'], status_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Статус обновлён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/client-statuses/<int:status_id>', methods=['DELETE'])
@admin_required
def ref_delete_client_status(status_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM Clients WHERE StatusID=?", status_id)
        cnt = cur.fetchone()[0]
        if cnt > 0:
            conn.close()
            return jsonify({'error': f'Нельзя удалить: используется в {cnt} клиентах'}), 409
        cur.execute("DELETE FROM ClientStatuses WHERE StatusID=?", status_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Статус удалён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/product-categories', methods=['GET'])
@login_required
def ref_get_product_categories():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT CategoryID, CategoryName FROM ProductCategories ORDER BY CategoryID")
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/product-categories', methods=['POST'])
@admin_required
def ref_create_product_category():
    data = request.get_json()
    if not data.get('name'):
        return jsonify({'error': 'Название обязательно'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO ProductCategories (CategoryName) OUTPUT INSERTED.CategoryID VALUES (?)", data['name'])
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'id': row[0], 'message': 'Категория создана'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/product-categories/<int:cat_id>', methods=['PUT'])
@admin_required
def ref_update_product_category(cat_id):
    data = request.get_json()
    if not data.get('name'):
        return jsonify({'error': 'Название обязательно'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("UPDATE ProductCategories SET CategoryName=? WHERE CategoryID=?", data['name'], cat_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Категория обновлена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/product-categories/<int:cat_id>', methods=['DELETE'])
@admin_required
def ref_delete_product_category(cat_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM Products WHERE CategoryID=?", cat_id)
        cnt = cur.fetchone()[0]
        if cnt > 0:
            conn.close()
            return jsonify({'error': f'Нельзя удалить: используется в {cnt} товарах'}), 409
        cur.execute("DELETE FROM ProductCategories WHERE CategoryID=?", cat_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Категория удалена'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/products', methods=['GET'])
@login_required
def ref_get_products():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT p.ProductID, p.ProductName, p.SKU, p.Unit, p.Price,
                   pc.CategoryID, pc.CategoryName, p.IsActive,
                   p.TileWidth, p.TileLength, p.TileThickness
            FROM Products p
            INNER JOIN ProductCategories pc ON p.CategoryID = pc.CategoryID
            ORDER BY pc.CategoryName, p.ProductName
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/products', methods=['POST'])
@admin_required
def ref_create_product():
    data = request.get_json()
    if not data.get('name') or not data.get('category_id') or not data.get('price'):
        return jsonify({'error': 'Название, категория и цена обязательны'}), 400
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO Products (ProductName, SKU, CategoryID, Unit, Price, IsActive, TileWidth, TileLength, TileThickness)
            OUTPUT INSERTED.ProductID
            VALUES (?, ?, ?, N'шт.', ?, 1, ?, ?, ?)
        """, data['name'], data.get('sku') or None, data['category_id'],
             float(data['price']),
             float(data.get('tile_width') or 0) or None,
             float(data.get('tile_length') or 0) or None,
             float(data.get('tile_thickness') or 0) or None)
        row = cur.fetchone()
        conn.commit()
        conn.close()
        return jsonify({'id': row[0], 'message': 'Товар создан'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/products/<int:product_id>', methods=['PUT'])
@admin_required
def ref_update_product(product_id):
    data = request.get_json()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE Products SET
                ProductName   = ISNULL(?, ProductName),
                SKU           = ?,
                CategoryID    = ISNULL(?, CategoryID),
                Price         = ISNULL(?, Price),
                IsActive      = ISNULL(?, IsActive),
                TileWidth     = ?,
                TileLength    = ?,
                TileThickness = ?
            WHERE ProductID = ?
        """,
            data.get('name'),
            data.get('sku') or None,
            data.get('category_id'),
            float(data['price']) if data.get('price') is not None else None,
            data.get('is_active'),
            float(data.get('tile_width') or 0) or None,
            float(data.get('tile_length') or 0) or None,
            float(data.get('tile_thickness') or 0) or None,
            product_id
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Товар обновлён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ref/products/<int:product_id>', methods=['DELETE'])
@admin_required
def ref_delete_product(product_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM DealItems WHERE ProductID=?", product_id)
        cnt1 = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM CommercialOfferItems WHERE ProductID=?", product_id)
        cnt2 = cur.fetchone()[0]
        total = cnt1 + cnt2
        if total > 0:
            conn.close()
            return jsonify({'error': f'Нельзя удалить: товар используется в {total} позициях (сделки: {cnt1}, КП: {cnt2})'}), 409
        cur.execute("DELETE FROM Products WHERE ProductID=?", product_id)
        conn.commit()
        conn.close()
        return jsonify({'message': 'Товар удалён'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# РЕЙТИНГ МЕНЕДЖЕРОВ
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/stats/ranking', methods=['GET'])
@login_required
def get_managers_ranking():
    period = request.args.get('period', 'month')
    now = datetime.now()
    try:
        conn = get_connection()
        cur = conn.cursor()

        if period == 'month':
            cur.execute("""
                SELECT u.UserID, u.FullName AS ManagerName,
                       ISNULL(SUM(d.Budget), 0) AS TotalSales,
                       COUNT(DISTINCT d.DealID) AS DealsCount
                FROM Users u
                INNER JOIN Roles r ON u.RoleID = r.RoleID AND r.RoleName = 'manager'
                LEFT JOIN Deals d ON d.ManagerID = u.UserID
                WHERE d.DealID IS NULL OR EXISTS (
                    SELECT 1 FROM DealStages ds 
                    WHERE ds.StageID = d.StageID AND ds.IsCompleted = 1
                    AND d.CompletedAt IS NOT NULL
                    AND YEAR(d.CompletedAt)=? AND MONTH(d.CompletedAt)=?
                )
                AND u.IsActive = 1
                GROUP BY u.UserID, u.FullName
                ORDER BY TotalSales DESC
            """, now.year, now.month)
        else:
            cur.execute("""
                SELECT u.UserID, u.FullName AS ManagerName,
                       ISNULL(SUM(d.Budget), 0) AS TotalSales,
                       COUNT(DISTINCT d.DealID) AS DealsCount
                FROM Users u
                INNER JOIN Roles r ON u.RoleID = r.RoleID AND r.RoleName = 'manager'
                LEFT JOIN Deals d ON d.ManagerID = u.UserID
                WHERE d.DealID IS NULL OR EXISTS (
                    SELECT 1 FROM DealStages ds 
                    WHERE ds.StageID = d.StageID AND ds.IsCompleted = 1
                )
                AND u.IsActive = 1
                GROUP BY u.UserID, u.FullName
                ORDER BY TotalSales DESC
            """)
        rows = cur.fetchall()
        conn.close()
        return jsonify(rows_to_list(cur, rows))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("=" * 50)
    print("  FSN GALLERY CRM — Запуск сервера")
    print("=" * 50)
    print(f"  Откройте в браузере: http://localhost:5000")
    print(f"  Для остановки нажмите Ctrl+C")
    print("=" * 50)
    app.run(debug=False, host='0.0.0.0', port=5000)
